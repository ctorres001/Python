########
# Indentifica las ventas anuladas sin compras posteriores y actualiza el archivo destino directamente.
########

import pandas as pd
import os
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import traceback

# === CONFIGURACIONES ===
print("🚀 Iniciando proceso simplificado de actualización...")

# Rutas de archivos - ACTUALIZADA
ruta_fuente = r"D:\FNB\Reportes\19. Reportes IBR\00. Estructura Reporte\Procesado\Archivo_Procesado.xlsx"
ruta_destino = r"D:\FNB\Reportes\19. Reportes IBR\11. Actualización Base Retenciones\Archivos\Consolidado ventas anuladas - Todos los canales.xlsx"
ruta_destinatarios = r"D:\FNB\Reportes\19. Reportes IBR\11. Actualización Base Retenciones\Destinatarios\Listado.xlsx"
firma_path = r"D:\FNB\Reportes\19. Reportes IBR\01. Pendientes de Entrega\Firma\Firma_resized.jpg"

# Configuraciones
fecha_inicio = datetime(2025, 9, 1)
motivos_excluir = [
    'PRUEBAS', 'POSIBLE FRAUDE', 'DUPLICADO POR SISTEMAS',
    'INCIDENCIA BIOMETRÍA - REGULARIZACIÓN DE VENTA'
]

columnas_exportar = [
    'RESPONSABLE DE VENTA', 'SEDE', 'ALIADO COMERCIAL', 'CUENTA CONTRATO',
    'CLIENTE', 'DNI', 'TELÉFONO', 'NÚMERO TELÉFONO OPCIONAL', 'CORREO',
    'Nro. PEDIDO VENTA', 'IMPORTE (S./)', 'Nro. DE CUOTAS', 'FECHA VENTA',
    'TIPO DESPACHO', 'ESTADO', 'FECHA ANULACIÓN', 'MOTIVO ANULACIÓN',
    'PRODUCTO_1', 'CANAL_VENTA'
]

def verificar_archivos():
    """Verifica que los archivos necesarios existan"""
    print("🔍 Verificando archivos...")
    
    if not os.path.exists(ruta_fuente):
        raise FileNotFoundError(f"❌ Archivo fuente no encontrado: {ruta_fuente}")
    print(f"✅ Archivo fuente: OK")
    
    if not os.path.exists(ruta_destino):
        raise FileNotFoundError(f"❌ Archivo destino no encontrado: {ruta_destino}")
    print(f"✅ Archivo destino: OK")
    
    return True

def procesar_datos_fuente():
    """Carga y procesa los datos del archivo fuente - MEJORADO para manejar fecha/hora"""
    print("📂 Cargando y procesando datos fuente...")
    
    # Cargar datos
    df = pd.read_excel(ruta_fuente, dtype=str)
    print(f"📊 Registros cargados: {len(df)}")
    
    # Limpiar espacios en blanco
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x) 
                  if col.dtype == 'object' else col)
    
    # Procesar fechas - MEJORADO para incluir hora si está disponible
    print("📅 Procesando fechas y horas...")
    
    # Verificar si existe columna de hora
    columna_hora = None
    if 'HORA VENTA' in df.columns:
        columna_hora = 'HORA VENTA'
    elif 'HORA' in df.columns:
        columna_hora = 'HORA'
    
    if columna_hora:
        print(f"✅ Encontrada columna de hora: {columna_hora}")
        # Combinar fecha y hora
        df['FECHA VENTA'] = pd.to_datetime(
            df['FECHA VENTA'].astype(str) + ' ' + df[columna_hora].astype(str), 
            errors='coerce'
        )
    else:
        print("⚠️ No se encontró columna de hora específica, usando solo fecha")
        df['FECHA VENTA'] = pd.to_datetime(df['FECHA VENTA'], errors='coerce')
    
    # Procesar fecha de anulación (probablemente sin hora)
    df['FECHA ANULACIÓN'] = pd.to_datetime(df['FECHA ANULACIÓN'], errors='coerce')
    
    # Mostrar algunas fechas de ejemplo para verificar
    print("📅 Ejemplos de fechas procesadas:")
    fechas_ejemplo = df['FECHA VENTA'].dropna().head(3)
    for fecha in fechas_ejemplo:
        print(f"    - {fecha}")
    
    # Filtrar por fecha (desde septiembre 2025)
    df = df[df['FECHA VENTA'] >= fecha_inicio]
    print(f"📅 Registros desde {fecha_inicio.strftime('%B %Y')}: {len(df)}")
    
    return df

def aplicar_filtros_anulaciones(df):
    """Aplica filtros para obtener solo las anulaciones válidas - CORREGIDO"""
    print("🔍 Aplicando filtros de anulaciones...")
    
    # Guardar copia completa para validación posterior
    df_completo = df.copy()
    
    # Mostrar estadísticas iniciales
    print(f"📊 Total registros: {len(df_completo)}")
    print(f"📊 Estados únicos: {df_completo['ESTADO'].unique()}")
    print(f"📊 Registros ANULADO: {len(df_completo[df_completo['ESTADO'] == 'ANULADO'])}")
    
    # Filtros principales para anulaciones
    filtro = (
        (df['ESTADO'] == 'ANULADO') &
        (df['ALIADO COMERCIAL'].str.upper() != 'CARDIF') &
        (~df['MOTIVO ANULACIÓN'].str.upper().isin([m.upper() for m in motivos_excluir]))
    )
    
    df_anuladas = df[filtro].copy()
    print(f"✅ Anulaciones válidas después de filtros: {len(df_anuladas)}")
    
    # Validar fechas de anulación
    sin_fecha_anulacion = df_anuladas['FECHA ANULACIÓN'].isna().sum()
    if sin_fecha_anulacion > 0:
        print(f"⚠️ {sin_fecha_anulacion} registros sin fecha de anulación")
    
    # Estados que NO consideramos como compras válidas posteriores
    estados_excluir = ['ANULADO', 'PENDIENTE DE ANULACIÓN']
    
    # Validar compras posteriores - CORREGIDO
    print("🔍 Validando compras posteriores con fecha y hora...")
    df_final = []
    excluidos_por_compras_posteriores = 0
    
    for idx, fila in df_anuladas.iterrows():
        cuenta = fila['CUENTA CONTRATO']
        fecha_hora_venta_anulada = fila['FECHA VENTA']  # Fecha/hora de la venta que fue anulada
        
        # Verificar que la fecha de venta no sea NaT
        if pd.isna(fecha_hora_venta_anulada):
            print(f"⚠️ Saltando registro sin fecha de venta: cuenta {cuenta}")
            continue
        
        # Buscar compras posteriores en la misma cuenta
        # CLAVE: Comparar con fecha/hora de VENTA (no de anulación)
        # y excluir estados ANULADO y PENDIENTE DE ANULACIÓN
        posteriores = df_completo[
            (df_completo['CUENTA CONTRATO'] == cuenta) &
            (df_completo['FECHA VENTA'] > fecha_hora_venta_anulada) &  # Posterior a la venta anulada
            (~df_completo['ESTADO'].isin(estados_excluir))  # NO anuladas ni pendientes de anulación
        ]
        
        # Debug: mostrar información para las primeras cuentas
        if not posteriores.empty and excluidos_por_compras_posteriores < 5:
            print(f"🔍 Cuenta {cuenta}: Encontradas {len(posteriores)} compras válidas posteriores a {fecha_hora_venta_anulada}")
            for _, post in posteriores.head(2).iterrows():  # Mostrar máximo 2 ejemplos
                print(f"    - Fecha: {post['FECHA VENTA']} | Estado: {post['ESTADO']} | Pedido: {post.get('Nro. PEDIDO VENTA', 'N/A')}")
        
        # Solo incluir si NO hay compras posteriores válidas
        if posteriores.empty:
            df_final.append(fila)
        else:
            excluidos_por_compras_posteriores += 1
    
    df_resultado = pd.DataFrame(df_final)
    print(f"✅ Anulaciones SIN compras posteriores: {len(df_resultado)}")
    print(f"❌ Excluidas por compras posteriores: {excluidos_por_compras_posteriores}")
    
    return df_resultado

def concatenar_productos(df):
    """Concatena múltiples columnas de productos"""
    print("🔧 Concatenando productos...")
    
    # Buscar columnas PRODUCTO_X
    columnas_producto = [col for col in df.columns if col.startswith('PRODUCTO_')]
    
    if not columnas_producto:
        print("⚠️ No se encontraron columnas de productos múltiples")
        if 'PRODUCTO' in df.columns:
            df['PRODUCTO_1'] = df['PRODUCTO']
        else:
            df['PRODUCTO_1'] = ''
        return df
    
    # Ordenar columnas numéricamente
    columnas_producto.sort(key=lambda x: int(x.split('_')[1]) if x.split('_')[1].isdigit() else 0)
    
    # Concatenar productos por fila
    productos_concatenados = []
    for _, fila in df.iterrows():
        productos = []
        for col in columnas_producto:
            valor = fila[col] if col in fila else ''
            if pd.notna(valor) and str(valor).strip() and str(valor).strip().upper() != 'NAN':
                productos.append(str(valor).strip())
        
        # Unir con separador
        resultado = ' | '.join(productos) if productos else ''
        productos_concatenados.append(resultado)
    
    df['PRODUCTO_1'] = productos_concatenados
    print(f"✅ Productos concatenados: {len([p for p in productos_concatenados if p])} registros con datos")
    
    return df

def formatear_datos(df):
    """Aplica formatos específicos a los datos - CORREGIDO para ordenar por FECHA ANULACIÓN"""
    print("🎨 Aplicando formatos...")
    
    # Formatear importes (2 decimales)
    def formatear_importe(valor):
        if pd.isna(valor) or valor == '':
            return ''
        try:
            numero = float(str(valor).replace(',', ''))
            return f"{numero:.2f}"
        except:
            return str(valor)
    
    df['IMPORTE (S./)'] = df['IMPORTE (S./)'].apply(formatear_importe)
    
    # CORREGIDO: Guardar fecha de anulación ANTES de formatear para ordenar correctamente
    fecha_anulacion_para_ordenar = df['FECHA ANULACIÓN'].copy()
    
    # Formatear fechas - PRESERVAR HORA SI EXISTE en FECHA VENTA
    def formatear_fecha_con_hora(fecha):
        if pd.isna(fecha):
            return ''
        # Si tiene hora diferente de 00:00, mostrar fecha y hora
        if fecha.hour != 0 or fecha.minute != 0 or fecha.second != 0:
            return fecha.strftime('%d/%m/%Y %H:%M:%S')
        else:
            return fecha.strftime('%d/%m/%Y')
    
    # Formatear fechas
    df['FECHA VENTA'] = df['FECHA VENTA'].apply(formatear_fecha_con_hora)
    df['FECHA ANULACIÓN'] = df['FECHA ANULACIÓN'].dt.strftime('%d/%m/%Y')
    
    # Ordenar por FECHA ANULACIÓN usando la fecha sin formatear
    df = df.iloc[fecha_anulacion_para_ordenar.argsort()].reset_index(drop=True)
    
    print("✅ Formatos aplicados y datos ordenados por FECHA ANULACIÓN")
    return df

def obtener_registros_nuevos(df_procesados):
    """Identifica qué registros son realmente nuevos"""
    print("🔍 Identificando registros nuevos...")
    
    try:
        # Cargar archivo destino
        df_existente = pd.read_excel(ruta_destino, sheet_name='Ventas anuladas')
        pedidos_existentes = set(df_existente['Nro. PEDIDO VENTA'].dropna().astype(str))
        
        print(f"📊 Registros en archivo destino: {len(df_existente)}")
        print(f"📊 Pedidos únicos existentes: {len(pedidos_existentes)}")
        
        # Filtrar solo nuevos
        df_nuevos = df_procesados[
            ~df_procesados['Nro. PEDIDO VENTA'].astype(str).isin(pedidos_existentes)
        ].copy()
        
        duplicados = len(df_procesados) - len(df_nuevos)
        print(f"📈 Total procesados: {len(df_procesados)}")
        print(f"🔄 Ya existentes: {duplicados}")
        print(f"✨ Nuevos a agregar: {len(df_nuevos)}")
        
        return df_nuevos, len(df_existente)
        
    except Exception as e:
        print(f"❌ Error al identificar registros nuevos: {e}")
        raise

def agregar_al_archivo_destino(df_nuevos):
    """Agrega los nuevos registros al archivo destino usando openpyxl"""
    if df_nuevos.empty:
        print("ℹ️ No hay registros nuevos para agregar")
        return True
    
    print(f"💾 Agregando {len(df_nuevos)} registros al archivo destino...")
    
    try:
        # Cargar workbook
        wb = load_workbook(ruta_destino)
        ws = wb['Ventas anuladas']
        
        # Encontrar última fila con datos
        ultima_fila = 1
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value:
                ultima_fila = fila
            else:
                break
        
        # Mapear columnas
        encabezados = [str(ws.cell(row=1, column=col).value).strip() 
                      for col in range(1, ws.max_column + 1)]
        mapeo_columnas = {enc: idx + 1 for idx, enc in enumerate(encabezados)}
        
        # Configurar formato
        font_aptos = Font(name='Aptos', size=8)
        alignment_left = Alignment(horizontal='left', vertical='center')
        
        # Agregar datos
        fila_actual = ultima_fila + 1
        for _, row in df_nuevos.iterrows():
            for columna in columnas_exportar:
                if columna in mapeo_columnas:
                    col_idx = mapeo_columnas[columna]
                    valor = row[columna] if columna in row else ""
                    valor = "" if pd.isna(valor) else valor
                    
                    cell = ws.cell(row=fila_actual, column=col_idx, value=valor)
                    cell.font = font_aptos
                    cell.alignment = alignment_left
            
            fila_actual += 1
        
        # Guardar
        wb.save(ruta_destino)
        wb.close()
        
        print(f"✅ {len(df_nuevos)} registros agregados exitosamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al agregar registros: {e}")
        traceback.print_exc()
        return False

def enviar_correo_resumen(exito, total_procesados=0, total_nuevos=0, error_msg=""):
    """Envía correo con resumen del proceso"""
    try:
        if not os.path.exists(ruta_destinatarios):
            print("❌ Archivo de destinatarios no encontrado")
            return
        
        df_dest = pd.read_excel(ruta_destinatarios)
        destinatarios = df_dest.iloc[:, 0].dropna().tolist()
        copia = df_dest.iloc[:, 1].dropna().tolist() if len(df_dest.columns) > 1 else []
        
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        
        mail.To = "; ".join(destinatarios)
        if copia:
            mail.CC = "; ".join(copia)
        
        fecha_proceso = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        if exito:
            asunto = f"✅ Actualización Base Retenciones - {fecha_proceso}"
            cuerpo = f"""<html><body style='font-family:Aptos, sans-serif; font-size:11pt;'>
Buenos días:<br><br>
Se completó exitosamente la actualización simplificada de la base de retenciones.<br><br>
<b>📊 Resumen:</b><br>
• Registros procesados: {total_procesados}<br>
• Registros nuevos agregados: {total_nuevos}<br>
• Proceso: Simplificado (sin archivos temporales)<br>
• Fecha/hora: {fecha_proceso}<br><br>
<b>✅ Validaciones aplicadas:</b><br>
• Filtrado por fecha (desde Sep 2025)<br>
• Solo anulaciones válidas<br>
• Sin compras posteriores (comparando fecha/hora de venta)<br>
• Productos concatenados<br>
• Ordenamiento por fecha de anulación<br><br>
Saludos cordiales.<br><br>
<img src="cid:firmaimg">
</body></html>"""
        else:
            asunto = f"❌ Error en Actualización - {fecha_proceso}"
            cuerpo = f"""<html><body style='font-family:Aptos, sans-serif; font-size:11pt;'>
Buenos días:<br><br>
Ocurrió un error durante la actualización:<br><br>
<b>❌ Error:</b><br>
{error_msg}<br><br>
<b>⏰ Fecha/hora:</b> {fecha_proceso}<br><br>
Por favor revisar.<br><br>
<img src="cid:firmaimg">
</body></html>"""
        
        mail.Subject = asunto
        mail.HTMLBody = cuerpo
        
        if os.path.exists(firma_path):
            attach = mail.Attachments.Add(firma_path)
            attach.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "firmaimg")
        
        mail.Display()
        print("📧 Correo preparado")
        
    except Exception as e:
        print(f"❌ Error en correo: {e}")

def debug_compras_posteriores_detallado(df, cuenta_ejemplo):
    """Función para debuggear una cuenta específica con fecha/hora"""
    print(f"\n🔍 DEBUG DETALLADO - Analizando cuenta: {cuenta_ejemplo}")
    
    # Filtrar por cuenta
    registros_cuenta = df[df['CUENTA CONTRATO'] == cuenta_ejemplo].copy()
    registros_cuenta = registros_cuenta.sort_values('FECHA VENTA')
    
    print(f"📊 Total registros para cuenta {cuenta_ejemplo}: {len(registros_cuenta)}")
    
    estados_excluir = ['ANULADO', 'PENDIENTE DE ANULACIÓN']
    
    for idx, reg in registros_cuenta.iterrows():
        fecha_venta = reg['FECHA VENTA']
        fecha_anulacion = reg['FECHA ANULACIÓN'] if pd.notna(reg['FECHA ANULACIÓN']) else 'Sin fecha'
        estado = reg['ESTADO']
        pedido = reg.get('Nro. PEDIDO VENTA', 'N/A')
        
        # Formato de fecha/hora
        if pd.notna(fecha_venta):
            fecha_str = fecha_venta.strftime('%d/%m/%Y %H:%M:%S') if fecha_venta.hour != 0 or fecha_venta.minute != 0 else fecha_venta.strftime('%d/%m/%Y')
        else:
            fecha_str = 'Sin fecha'
        
        print(f"  • Pedido: {pedido} | Venta: {fecha_str} | Estado: {estado}")
        if estado == 'ANULADO':
            print(f"    └── Anulación: {fecha_anulacion}")
            
            # Si es anulado, buscar compras posteriores
            if pd.notna(fecha_venta):
                posteriores = registros_cuenta[
                    (registros_cuenta['FECHA VENTA'] > fecha_venta) &
                    (~registros_cuenta['ESTADO'].isin(estados_excluir))
                ]
                if not posteriores.empty:
                    print(f"    └── ❌ TIENE {len(posteriores)} compras posteriores válidas:")
                    for _, post in posteriores.iterrows():
                        post_fecha_str = post['FECHA VENTA'].strftime('%d/%m/%Y %H:%M:%S') if post['FECHA VENTA'].hour != 0 else post['FECHA VENTA'].strftime('%d/%m/%Y')
                        print(f"        • {post_fecha_str} - {post['ESTADO']} (Pedido: {post.get('Nro. PEDIDO VENTA', 'N/A')})")
                else:
                    print(f"    └── ✅ NO tiene compras posteriores válidas")
    
    return registros_cuenta

def main():
    """Proceso principal simplificado - CORREGIDO"""
    try:
        # 1. Verificar archivos
        verificar_archivos()
        
        # 2. Procesar datos fuente
        df_fuente = procesar_datos_fuente()
        
        if df_fuente.empty:
            print("⚠️ No hay datos para procesar")
            enviar_correo_resumen(True, 0, 0)
            return
        
        # DEBUG OPCIONAL: Uncomment para debuggear una cuenta específica
        # debug_compras_posteriores_detallado(df_fuente, "TU_CUENTA_EJEMPLO")
        
        # 3. Aplicar filtros
        df_anuladas = aplicar_filtros_anulaciones(df_fuente)
        
        if df_anuladas.empty:
            print("⚠️ No hay anulaciones válidas")
            enviar_correo_resumen(True, 0, 0)
            return
        
        # 4. Procesar productos y formatear
        df_anuladas = concatenar_productos(df_anuladas)
        df_final = formatear_datos(df_anuladas)
        
        # 5. Preparar datos para exportar
        df_exportar = df_final[columnas_exportar].copy()
        
        # 6. Identificar registros nuevos
        df_nuevos, registros_originales = obtener_registros_nuevos(df_exportar)
        
        # 7. Agregar al archivo destino
        if not df_nuevos.empty:
            exito = agregar_al_archivo_destino(df_nuevos)
            
            if exito:
                print("🎉 Proceso completado exitosamente!")
                print(f"📊 Resumen:")
                print(f"   • Registros base: {registros_originales}")
                print(f"   • Nuevos agregados: {len(df_nuevos)}")
                print(f"   • Total final: {registros_originales + len(df_nuevos)}")
                
                enviar_correo_resumen(True, len(df_exportar), len(df_nuevos))
            else:
                enviar_correo_resumen(False, error_msg="Error al agregar registros al archivo destino")
        else:
            print("ℹ️ No hay registros nuevos para procesar")
            enviar_correo_resumen(True, len(df_exportar), 0)
            
    except Exception as e:
        error_msg = f"Error: {str(e)}\n\nDetalle:\n{traceback.format_exc()}"
        print(f"❌ Error en proceso principal: {e}")
        print(traceback.format_exc())
        enviar_correo_resumen(False, error_msg=error_msg)

if __name__ == "__main__":
    main()
    print("🏁 Proceso finalizado.")