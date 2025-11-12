import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Rutas
archivo_origen = r"D:\FNB\Reportes\11. Reporte Clausulas Acuerdo Comercial FNB - Penalidades\Bitacora Penalidades FNB - OCTUBRE 2025.xlsx"
carpeta_destino = r"D:\FNB\Reportes\11. Reporte Clausulas Acuerdo Comercial FNB - Penalidades\2025\10 Octubre 2025"

# Crear carpeta destino si no existe
os.makedirs(carpeta_destino, exist_ok=True)

def aplicar_formato(archivo_excel):
    """Aplica formato al archivo Excel generado"""
    wb = load_workbook(archivo_excel)
    ws = wb.active
    
    # Configurar fuente Aptos, tamaño 8
    fuente_normal = Font(name='Aptos Narrow', size=8)
    fuente_encabezado = Font(name='Aptos Narrow', size=8, color='FFFFFF', bold=True)
    
    # Relleno negro para encabezado
    relleno_negro = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    
    # Alineación centrada
    alineacion = Alignment(horizontal='center', vertical='center')
    
    # Aplicar formato a encabezado (fila 1)
    for cell in ws[1]:
        cell.font = fuente_encabezado
        cell.fill = relleno_negro
        cell.alignment = alineacion
    
    # Alto de fila para encabezado
    ws.row_dimensions[1].height = 11.25
    
    # Aplicar formato al resto de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = fuente_normal
            cell.alignment = alineacion
            
            # Formato de fecha para columnas que contienen fechas
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY'
        
        # Alto de fila
        ws.row_dimensions[cell.row].height = 11.25
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(archivo_excel)
    wb.close()

# Procesar hoja "BD Colocaciones FNB"
print("Procesando hoja: BD Colocaciones FNB")
try:
    df_colocaciones = pd.read_excel(archivo_origen, sheet_name='BD Colocaciones FNB')
    
    # Obtener proveedores únicos
    proveedores_col = df_colocaciones['Nombre de Proveedor'].dropna().unique()
    
    for proveedor in proveedores_col:
        # Filtrar datos por proveedor
        df_filtrado = df_colocaciones[df_colocaciones['Nombre de Proveedor'] == proveedor]
        
        # Nombre del archivo
        nombre_archivo = f"No cumplir con la entrega del producto al cliente en plazo {proveedor}.xlsx"
        ruta_completa = os.path.join(carpeta_destino, nombre_archivo)
        
        # Guardar archivo
        df_filtrado.to_excel(ruta_completa, index=False, engine='openpyxl')
        
        # Aplicar formato
        aplicar_formato(ruta_completa)
        
        print(f"  ✓ Creado: {nombre_archivo}")
    
    print(f"Total archivos creados (BD Colocaciones): {len(proveedores_col)}")
    
except Exception as e:
    print(f"Error procesando BD Colocaciones FNB: {str(e)}")

# Procesar hoja "Bandeja Anulacion"
print("\nProcesando hoja: Bandeja Anulacion")
try:
    df_anulacion = pd.read_excel(archivo_origen, sheet_name='Bandeja Anulacion')
    
    # Obtener aliados únicos
    aliados = df_anulacion['ALIADO COMERCIAL'].dropna().unique()
    
    for aliado in aliados:
        # Filtrar datos por aliado
        df_filtrado = df_anulacion[df_anulacion['ALIADO COMERCIAL'] == aliado]
        
        # Nombre del archivo
        nombre_archivo = f"No contar con stock de un determinado producto {aliado}.xlsx"
        ruta_completa = os.path.join(carpeta_destino, nombre_archivo)
        
        # Guardar archivo
        df_filtrado.to_excel(ruta_completa, index=False, engine='openpyxl')
        
        # Aplicar formato
        aplicar_formato(ruta_completa)
        
        print(f"  ✓ Creado: {nombre_archivo}")
    
    print(f"Total archivos creados (Bandeja Anulacion): {len(aliados)}")
    
except Exception as e:
    print(f"Error procesando Bandeja Anulacion: {str(e)}")

print("\n¡Proceso completado exitosamente!")