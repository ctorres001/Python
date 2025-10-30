import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from datetime import datetime, timedelta
import win32com.client as win32
import time
import sys
from pathlib import Path
import logging
from typing import List, Tuple, Optional
import dataframe_image as dfi
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import psutil

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ReporteFNB:
    def __init__(self):
        # NOTA: La fecha se ajustará después de seleccionar la actividad
        self.fecha_filtro = None  # Se define en seleccionar_actividad()
        self.fecha_str = None
        self.fecha_mostrar = None
        
        self.ruta_base = Path(r'D:\FNB\Reportes\19. Reportes IBR\02. Avance Colocaciones')
        self.ruta_salida = self.ruta_base / 'Bases'
        self.ruta_imagenes = self.ruta_base / 'Imagenes'
        self.tipo_actividad = None
        self.df = None
        self.responsables = None
        self.proveedores = None
        self.solo_canales_propios = False

        # NUEVOS PARÁMETROS
        self.modo_render = "chrome"
        self.formatear_excel = True

        # Crear carpetas si no existen
        self.ruta_salida.mkdir(exist_ok=True)
        self.ruta_imagenes.mkdir(exist_ok=True)

        # Columnas necesarias
        self.columnas_deseadas = [
            'F. Registro', 'F. Entrega', 'Cuenta Contrato', 'DNI', 'Nombre y Apellido de Cliente',
            'Teléfono', 'Correo Electronico', 'Distrito', 'NSE', 'N° de Contrato', 'N° de Boleta',
            'Pedido Venta', 'Importe Colocación  S/', 'Importe Financiamiento  S/', 'N° de Cuotas',
            'Nombre Responsable de Venta', 'Nombre de Proveedor', 'Nombre Tienda de Venta',
            'Modalidad de Entrega', 'Estado de Entrega', 'PRODUCTO 1', 'SKU 1', 'PRODUCTO 2',
            'SKU 2', 'PRODUCTO 3', 'SKU 3', 'PRODUCTO 4', 'SKU 4', 'Asesor', 'Tiempo de Entrega',
            'Rangos', 'Zona de Venta', 'Marca', 'Modelo', 'Canal', 'Tipo de Producto',
            'Tipo Instalación', '# Transacciones'
        ]

    def seleccionar_archivo(self) -> str:
        """Selecciona el archivo Excel de origen"""
        root = tk.Tk()
        root.withdraw()
        archivo_origen = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")]
        )
        root.destroy()
        return archivo_origen

    def seleccionar_actividad(self) -> str:
        """Selecciona el tipo de actividad y configura la fecha según el tipo"""
        while True:
            opcion = input("Selecciona una opción:\n1. Avance de Ventas\n2. Cierre de Ventas\nOpción: ").strip()
            if opcion == '1':
                self.tipo_actividad = "Avance de Ventas FNB"
                # Para Avance: usa día anterior
                self.fecha_filtro = datetime.now() - timedelta(days=1)
                break
            elif opcion == '2':
                self.tipo_actividad = "Cierre de Ventas FNB"
                # Para Cierre: usa último día del mes anterior
                hoy = datetime.now()
                primer_dia_mes_actual = hoy.replace(day=1)
                self.fecha_filtro = primer_dia_mes_actual - timedelta(days=1)
                break
            else:
                print("Opción inválida. Intenta de nuevo.")
        
        # Configurar formatos de fecha
        self.fecha_str = self.fecha_filtro.strftime('%Y-%m-%d')
        self.fecha_mostrar = self.fecha_filtro.strftime('%d/%m/%Y')

        while True:
            opcion = input("¿Deseas procesar solo Canales Propios?\n1. Sí (IBR PERU y SALESLAND)\n2. No (Canales Completos)\nOpción: ").strip()
            if opcion == '1':
                self.solo_canales_propios = True
                break
            elif opcion == '2':
                self.solo_canales_propios = False
                break
            else:
                print("Opción inválida. Intenta de nuevo.")

        return self.tipo_actividad
    
    def cargar_datos(self, archivo_origen: str) -> pd.DataFrame:
        """Carga y procesa los datos del archivo Excel"""
        try:
            logger.info(f"Cargando datos desde: {archivo_origen}")
            df = pd.read_excel(archivo_origen, sheet_name='BD Colocaciones FNB', header=0, dtype=str)

            # Limpieza de columnas y contenido
            df.columns = df.columns.str.strip().str.replace('\n', ' ', regex=True)
            df = df.astype(str).apply(lambda col: col.str.strip().str.replace('\n', ' ', regex=True))

            # Filtrar columnas necesarias
            columnas_existentes = [col for col in self.columnas_deseadas if col in df.columns]
            df = df[columnas_existentes]

            # Procesar fechas y filtrar por mes/año
            df['F. Entrega'] = pd.to_datetime(df['F. Entrega'], errors='coerce')
            mes_filtro = self.fecha_filtro.month
            anio_filtro = self.fecha_filtro.year

            df = df[
                (df['F. Entrega'].dt.month == mes_filtro) &
                (df['F. Entrega'].dt.year == anio_filtro) |
                (df['F. Entrega'].isna())
            ]

            # Convertir columnas numéricas
            df['Importe Colocación  S/'] = pd.to_numeric(df['Importe Colocación  S/'], errors='coerce').fillna(0)
            df['# Transacciones'] = pd.to_numeric(df['# Transacciones'], errors='coerce').fillna(0)

            logger.info(f"Datos cargados exitosamente: {len(df)} registros")
            return df

        except Exception as e:
            logger.error(f"Error cargando datos: {e}")
            raise

    def cargar_destinatarios(self):
        """Carga los archivos de destinatarios"""
        try:
            ruta_destinatarios = self.ruta_base / 'Destinatarios'

            self.responsables = pd.read_excel(ruta_destinatarios / 'Listado de correos Responsables.xlsx')
            self.proveedores = pd.read_excel(ruta_destinatarios / 'Listado de correos Proveedores.xlsx')

            # Limpiar nombres de columnas
            self.responsables.columns = self.responsables.columns.str.strip().str.replace('\n', ' ', regex=True)
            self.proveedores.columns = self.proveedores.columns.str.strip().str.replace('\n', ' ', regex=True)

            logger.info("Destinatarios cargados exitosamente")

        except Exception as e:
            logger.error(f"Error cargando destinatarios: {e}")
            raise
    def crear_tabla_dinamica(
            self,
            data: pd.DataFrame,
            filtro_col: str = None,
            filtro_val: str = None,
            tipo_tabla: str = 'general',
            omitir_subtotales: bool = False
    ) -> pd.DataFrame:
        """Crea una tabla dinámica optimizada"""
        try:
            if filtro_col and filtro_val:
                data_filtrada = data[data[filtro_col] == filtro_val].copy()
            else:
                data_filtrada = data.copy()

            if data_filtrada.empty:
                return pd.DataFrame()

            columna_mapping = {
                'canal': ('Canal', 'Canal de Venta'),
                'sede': ('Nombre Tienda de Venta', 'Sede'),
                'proveedor': ('Nombre de Proveedor', 'Proveedor')
            }

            col_secundaria, nombre_cabecera = columna_mapping.get(tipo_tabla, ('Nombre de Proveedor', 'Proveedor'))

            if col_secundaria not in data_filtrada.columns:
                return pd.DataFrame()

            filas_tabla = []

            if omitir_subtotales:
                agrupado = data_filtrada.groupby(col_secundaria).agg({
                    'Importe Colocación  S/': 'sum',
                    '# Transacciones': 'sum'
                }).reset_index().sort_values('Importe Colocación  S/', ascending=False)

                for _, row in agrupado.iterrows():
                    filas_tabla.append({
                        nombre_cabecera: row[col_secundaria],
                        'Importe S/': row['Importe Colocación  S/'],
                        '# Transacciones': row['# Transacciones'],
                        'Es_Subtotal': False
                    })

            else:
                estados_entrega = ['Producto Entregado', 'Pendiente de Entrega']
                for estado in estados_entrega:
                    if estado in data_filtrada['Estado de Entrega'].values:
                        estado_data = data_filtrada[data_filtrada['Estado de Entrega'] == estado]
                        subtotal_importe = estado_data['Importe Colocación  S/'].sum()
                        subtotal_transacciones = estado_data['# Transacciones'].sum()

                        filas_tabla.append({
                            nombre_cabecera: estado,
                            'Importe S/': subtotal_importe,
                            '# Transacciones': subtotal_transacciones,
                            'Es_Subtotal': True
                        })

                        agrupado = estado_data.groupby(col_secundaria).agg({
                            'Importe Colocación  S/': 'sum',
                            '# Transacciones': 'sum'
                        }).reset_index().sort_values('Importe Colocación  S/', ascending=False)

                        for _, row in agrupado.iterrows():
                            filas_tabla.append({
                                nombre_cabecera: row[col_secundaria],
                                'Importe S/': row['Importe Colocación  S/'],
                                '# Transacciones': row['# Transacciones'],
                                'Es_Subtotal': False
                            })

            total_importe = data_filtrada['Importe Colocación  S/'].sum()
            total_transacciones = data_filtrada['# Transacciones'].sum()

            filas_tabla.append({
                nombre_cabecera: "Total General",
                'Importe S/': total_importe,
                '# Transacciones': total_transacciones,
                'Es_Subtotal': True
            })

            df_tabla = pd.DataFrame(filas_tabla)

            df_tabla['% Participación'] = (
                (df_tabla['Importe S/'] / total_importe * 100).round(1) if total_importe > 0 else 0
            )

            df_final = df_tabla.copy()
            df_final['Importe S/'] = df_final['Importe S/'].apply(lambda x: f"S/ {x:,.0f}")
            df_final['# Transacciones'] = df_final['# Transacciones'].apply(lambda x: f"{x:,.0f}")
            df_final['% Participación'] = df_final['% Participación'].apply(
                lambda x: f"{x:.1f}%" if isinstance(x, float) else "0.0%")

            return df_final

        except Exception as e:
            logger.error(f"Error creando tabla dinámica: {e}")
            return pd.DataFrame()
    def crear_imagen_tabla(self, df_tabla: pd.DataFrame, nombre_archivo: str) -> Optional[str]:
        """Crea imagen de tabla completa usando dataframe_image y chrome sin recortes"""
        try:
            if df_tabla.empty:
                return None

            df_imagen = df_tabla.drop(columns=['Es_Subtotal'], errors='ignore').reset_index(drop=True)

            def aplicar_estilo(row):
                if row.name < len(df_tabla) and 'Es_Subtotal' in df_tabla.columns:
                    if df_tabla.iloc[row.name]['Es_Subtotal']:
                        return ['background-color: #D9D9D9; font-weight: bold'] * len(row)
                return ['background-color: white'] * len(row)

            df_styled = df_imagen.style.apply(aplicar_estilo, axis=1)

            df_styled = df_styled.set_table_styles([
                {'selector': 'th', 'props': [
                    ('background-color', '#000000'),
                    ('color', 'white'),
                    ('font-weight', 'bold'),
                    ('text-align', 'center'),
                    ('border', '1px solid #000'),
                    ('padding', '6px'),
                    ('font-family', 'Verdana'),
                    ('font-size', '8pt')
                ]},
                {'selector': 'tbody td:first-child', 'props': [
                    ('text-align', 'left'),
                    ('border', '1px solid #000'),
                    ('padding', '6px'),
                    ('font-family', 'Verdana'),
                    ('font-size', '8pt')
                ]},
                {'selector': 'tbody td:not(:first-child)', 'props': [
                    ('text-align', 'center'),
                    ('border', '1px solid #000'),
                    ('padding', '6px'),
                    ('font-family', 'Verdana'),
                    ('font-size', '8pt')
                ]},
                {'selector': 'table', 'props': [
                    ('border-collapse', 'collapse'),
                    ('border', '2px solid #000'),
                    ('width', '100%'),
                    ('max-height', 'none'),
                    ('overflow', 'visible')
                ]}
            ])

            df_styled = df_styled.hide(axis='index')

            ruta_imagen = self.ruta_imagenes / f"{nombre_archivo}.png"

            # Exportar tabla sin límites de scroll ni altura
            dfi.export(
                df_styled,
                str(ruta_imagen),
                fontsize=8,
                max_rows=-1,
                max_cols=-1,
                table_conversion="chrome",
                chrome_path=None
            )

            logger.info(f"Imagen creada completa: {ruta_imagen}")
            return str(ruta_imagen)

        except Exception as e:
            logger.error(f"Error creando imagen sin recorte: {e}")
            return None
    def generar_imagenes_proveedor(self, data: pd.DataFrame, proveedor: str) -> List[Tuple[str, str]]:
        """Genera las imágenes específicas según el proveedor"""
        imagenes = []
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        configuraciones = {
            "IBR PERU": [
                ('canal_venta', None, None, 'canal'),
                ('motos', 'Canal', 'MOTOS', 'proveedor'),
                ('materiales', 'Canal', 'FERRETERAS', 'proveedor'),
                ('digital', 'Canal', 'DIGITAL', 'proveedor')  # ← NUEVA LÍNEA
            ],
            "SALESLAND": [
                ('por_proveedor', None, None, 'proveedor'),
                ('por_sede', None, None, 'sede')
            ]
        }

        config = configuraciones.get(proveedor, [('general', None, None, 'canal')])

        for tipo, filtro_col, filtro_val, tipo_tabla in config:
            tabla = self.crear_tabla_dinamica(data, filtro_col, filtro_val, tipo_tabla)
            if not tabla.empty:
                nombre_archivo = f"{tipo}_{proveedor}_{timestamp}"
                ruta_imagen = self.crear_imagen_tabla(tabla, nombre_archivo)
                if ruta_imagen:
                    imagenes.append((tipo, ruta_imagen))

        # Inserta las 2 imágenes específicas para 'ALO CÁLIDDA' si es IBR PERU
        if proveedor == "IBR PERU":
            canal_data = data[data['Canal'] == 'ALO CÁLIDDA'].copy()

            if not canal_data.empty:
                for estado, tipo_suffix in [
                    ('Producto Entregado', 'alo_calidda_entregados'),
                    ('Pendiente de Entrega', 'alo_calidda_pendientes')
                ]:
                    df_estado = canal_data[canal_data['Estado de Entrega'] == estado]
                    if not df_estado.empty:
                        tabla_estado = self.crear_tabla_dinamica(df_estado, tipo_tabla='proveedor',
                                                                 omitir_subtotales=True)
                        if not tabla_estado.empty:
                            nombre_archivo = f"{tipo_suffix}_{proveedor}_{timestamp}"
                            ruta_imagen = self.crear_imagen_tabla(tabla_estado, nombre_archivo)
                            if ruta_imagen:
                                imagenes.append((tipo_suffix, ruta_imagen))

        return imagenes
    def guardar_excel_openpyxl(self, data: pd.DataFrame, ruta_archivo: str) -> bool:
        """Guarda el archivo Excel usando openpyxl con formato opcional"""
        try:
            data_clean = data.replace(['nan', 'NaN', pd.NA], '').fillna('')

            if 'F. Registro' in data_clean.columns:
                data_clean['F. Registro'] = pd.to_datetime(data_clean['F. Registro'], errors='coerce').dt.date
            if 'F. Entrega' in data_clean.columns:
                data_clean['F. Entrega'] = pd.to_datetime(data_clean['F. Entrega'], errors='coerce').dt.date

            if os.path.exists(ruta_archivo):
                self._cerrar_excel_procesos()

            with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                data_clean.to_excel(writer, index=False, sheet_name='Datos')

                if not self.formatear_excel:
                    return True  # Salta formato si se desactiva

                workbook = writer.book
                worksheet = workbook['Datos']

                header_font = Font(name='Aptos', size=8, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                data_font = Font(name='Aptos', size=8)

                for col in range(1, len(data_clean.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                for row in range(2, len(data_clean) + 2):
                    worksheet.row_dimensions[row].height = 11.25
                    for col in range(1, len(data_clean.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        if col <= 2:
                            cell.number_format = 'dd/mm/yyyy'

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"Archivo Excel guardado: {ruta_archivo}")
            return True

        except Exception as e:
            logger.error(f"Error guardando Excel: {e}")
            return False

    def _cerrar_excel_procesos(self):
        """Cierra procesos de Excel que bloqueen archivos"""
        try:
            for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                if proc.info['name'] and 'excel' in proc.info['name'].lower():
                    try:
                        if proc.open_files():
                            for f in proc.open_files():
                                if f.path.startswith(str(self.ruta_salida)) or f.path.endswith(".xlsx"):
                                    proc.terminate()
                                    time.sleep(1)
                                    break
                    except:
                        continue
        except:
            pass
    def enviar_correo(self, proveedor: str, ruta_archivo: str, imagenes: List[Tuple[str, str]]):
        """Envía el correo con archivo adjunto e imágenes embebidas"""
        try:
            fila = None
            if proveedor in self.responsables['Nombre Responsable de Venta'].values:
                fila = self.responsables[self.responsables['Nombre Responsable de Venta'] == proveedor].iloc[0]
            elif proveedor in self.proveedores['Nombre de Proveedor'].values:
                fila = self.proveedores[self.proveedores['Nombre de Proveedor'] == proveedor].iloc[0]

            if fila is None:
                logger.warning(f"Destinatario no encontrado: {proveedor}")
                return

            to = fila.iloc[1]
            cc = fila.iloc[2] if len(fila) > 2 else ""
            asunto = f"{self.tipo_actividad} - {proveedor} - {self.fecha_str}"

            outlook = win32.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = to
            mail.CC = cc
            mail.Subject = asunto
            mail.BodyFormat = 2  # HTML

            mail.Attachments.Add(ruta_archivo)

            for tipo, ruta_img in imagenes:
                if os.path.exists(ruta_img):
                    attachment = mail.Attachments.Add(ruta_img)
                    attachment.PropertyAccessor.SetProperty(
                        "http://schemas.microsoft.com/mapi/proptag/0x3712001E", tipo
                    )

            mail.Display()  # conserva la firma
            signature = mail.HTMLBody
            cuerpo = self.generar_cuerpo_correo(proveedor, imagenes)
            mail.HTMLBody = cuerpo + signature

            mail.Send()
            logger.info(f"Correo enviado a: {to} | CC: {cc}")

            self._limpiar_imagenes_temporales(imagenes)

        except Exception as e:
            logger.error(f"Error enviando correo para {proveedor}: {e}")

    def generar_cuerpo_correo(self, proveedor: str, imagenes: List[Tuple[str, str]]) -> str:
        """Genera el cuerpo del correo con imágenes embebidas"""
        base_texto = f"""Buenos días:<br><br>
        Se adjunta la BD de financiamientos FNB registrados al <strong>{self.fecha_mostrar}</strong>.<br><br>"""

        extra = ""

        if proveedor == "IBR PERU":
            extra += "<strong>Por Canal de Venta:</strong><br><br>"
            for tipo in ['canal_venta']:
                for img_tipo, ruta_img in imagenes:
                    if img_tipo == tipo:
                        extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

            extra += "<strong>Detalle Canal Aló Cálidda:</strong><br><br>"
            etiquetas = {
                'alo_calidda_entregados': "Productos Entregados<br>",
                'alo_calidda_pendientes': "Productos Pendientes de Entrega<br>"
            }

            for tipo in ['alo_calidda_entregados', 'alo_calidda_pendientes']:
                if tipo in etiquetas:
                    extra += f"{etiquetas[tipo]}<br>"
                    for img_tipo, ruta_img in imagenes:
                        if img_tipo == tipo:
                            extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

            extra += "<strong>Detalle Canal Motos:</strong><br><br>"
            for img_tipo, ruta_img in imagenes:
                if img_tipo == "motos":
                    extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

            extra += "<strong>Detalle Canal Materiales:</strong><br><br>"
            for img_tipo, ruta_img in imagenes:
                if img_tipo == "materiales":
                    extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

            # ← AGREGAR ESTA SECCIÓN
            extra += "<strong>Detalle Canal Digital:</strong><br><br>"
            for img_tipo, ruta_img in imagenes:
                if img_tipo == "digital":
                    extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

        elif proveedor == "SALESLAND":
            extra += "<strong>Estado de ventas por proveedor:</strong><br><br>"
            for img_tipo, ruta_img in imagenes:
                if img_tipo == "por_proveedor":
                    extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'
            extra += "<strong>Estado de ventas por sede:</strong><br><br>"
            for img_tipo, ruta_img in imagenes:
                if img_tipo == "por_sede":
                    extra += f'<img src="cid:{img_tipo}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

        else:
            extra += "<strong>Detalle de resultados:</strong><br><br>"
            for _, ruta_img in imagenes:
                extra += f'<img src="cid:{_}" style="width: 12cm; height: auto; max-width: 100%;"><br><br>'

        cierre = "Quedo atento a cualquier consulta,"
        return f'<div style="font-family:Aptos; font-size:11pt">{base_texto}{extra}{cierre}</div>'
    def _limpiar_imagenes_temporales(self, imagenes: List[Tuple[str, str]]):
        """Limpia archivos de imagen temporales"""
        for _, ruta_img in imagenes:
            try:
                if os.path.exists(ruta_img):
                    os.remove(ruta_img)
            except:
                pass

    def obtener_filtros_aplicados(self) -> List[Tuple[pd.DataFrame, str]]:
        """Ordena proveedores a procesar: IBR, SALESLAND, luego el resto (según opción)"""
        filtros_aplicados = []
        orden_prioridad = ["IBR PERU", "SALESLAND"]

        # Agregar canales propios
        for nombre in orden_prioridad:
            if nombre in self.responsables['Nombre Responsable de Venta'].values:
                filtro = self.df[self.df['Nombre Responsable de Venta'] == nombre]
                if not filtro.empty:
                    filtros_aplicados.append((filtro, nombre))
            elif nombre in self.proveedores['Nombre de Proveedor'].values:
                filtro = self.df[self.df['Nombre de Proveedor'] == nombre]
                if not filtro.empty:
                    filtros_aplicados.append((filtro, nombre))

        # Agregar el resto solo si aplica
        if not self.solo_canales_propios:
            for _, row in self.proveedores.iterrows():
                nombre = row['Nombre de Proveedor']
                if nombre in orden_prioridad:
                    continue
                if nombre == 'INTEGRA RETAIL S.A.C.':
                    filtro = self.df[
                        (self.df['Nombre de Proveedor'] == nombre) &
                        (self.df['Nombre Responsable de Venta'] == 'IBR PERU')
                    ]
                else:
                    filtro = self.df[self.df['Nombre de Proveedor'] == nombre]
                if not filtro.empty:
                    filtros_aplicados.append((filtro, nombre))

            for _, row in self.responsables.iterrows():
                nombre = row['Nombre Responsable de Venta']
                if nombre in orden_prioridad:
                    continue
                filtro = self.df[self.df['Nombre Responsable de Venta'] == nombre]
                if not filtro.empty:
                    filtros_aplicados.append((filtro, nombre))

        return filtros_aplicados

    def procesar_reporte(self):
        """Ejecución completa del proceso de reportes y correos"""
        try:
            archivo_origen = self.seleccionar_archivo()
            if not archivo_origen:
                logger.error("No se seleccionó archivo")
                return

            self.tipo_actividad = self.seleccionar_actividad()
            self.df = self.cargar_datos(archivo_origen)
            self.cargar_destinatarios()
            filtros_aplicados = self.obtener_filtros_aplicados()

            inicio_total = time.time()
            for data, proveedor in filtros_aplicados:
                try:
                    logger.info(f"Procesando: {proveedor}")
                    inicio = time.time()

                    imagenes = self.generar_imagenes_proveedor(data, proveedor)
                    nombre_archivo = f"{self.tipo_actividad} - {proveedor} - {self.fecha_str}.xlsx"
                    ruta_archivo = self.ruta_salida / nombre_archivo

                    if self.guardar_excel_openpyxl(data, str(ruta_archivo)):
                        self.enviar_correo(proveedor, str(ruta_archivo), imagenes)

                    duracion = round(time.time() - inicio, 2)
                    logger.info(f"✓ Terminado {proveedor} en {duracion} segundos\n")

                except Exception as e:
                    logger.error(f"Error procesando {proveedor}: {e}")
                    continue

            total = round(time.time() - inicio_total, 2)
            logger.info(f"✅ Proceso finalizado en {total} segundos")

        except Exception as e:
            logger.error(f"Error en proceso principal: {e}")
            raise
def main():
    try:
        reporte = ReporteFNB()
        reporte.procesar_reporte()
    except Exception as e:
        logger.error(f"Error fatal: {e}")
        input("Presiona Enter para salir...")

if __name__ == "__main__":
    main()
