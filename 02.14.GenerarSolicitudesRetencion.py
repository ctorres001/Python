########
# Identifica las ventas Pendientes de Anulación y actualiza el archivo destino.
########

import pandas as pd
import os
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import traceback

# === CONFIGURACIONES - SOLICITUDES DE ANULACIÓN ===
print("🚀 Iniciando proceso simplificado de solicitudes de anulación...")

# Rutas de archivos - ACTUALIZADAS
ruta_fuente_principal = r"D:\FNB\Reportes\19. Reportes IBR\11. Actualización Base Retenciones\Base Solicitudes Anulacion\Base Solicitudes Anulacion.xlsx"
ruta_fuente_complementaria = r"D:\FNB\Reportes\19. Reportes IBR\00. Estructura Reporte\Procesado\Archivo_Procesado.xlsx"
ruta_destino = r"D:\FNB\Reportes\19. Reportes IBR\11. Actualización Base Retenciones\Archivos\Solicitudes de anulaciones - Todos los canales.xlsx"
ruta_destinatarios = r"D:\FNB\Reportes\19. Reportes IBR\11. Actualización Base Retenciones\Destinatarios\Listado.xlsx"
firma_path = r"D:\FNB\Reportes\19. Reportes IBR\01. Pendientes de Entrega\Firma\Firma_resized.jpg"

# Configuraciones
fecha_inicio = datetime(2025, 9, 1)
motivos_excluir = [
    'PRUEBAS', 'POSIBLE FRAUDE', 'DUPLICADO POR SISTEMAS',
    'INCIDENCIA BIOMETRÍA - REGULARIZACIÓN DE VENTA'
]

estados_anulacion_validos = [
    'ANULACIÓN RECHAZADA', 'DERIVADO AL RESPONSABLE DE VENTA',
    'ASIGNADO A BO', 'POR DERIVAR AL RESPONSABLE DE LA VENTA'
]

columnas_exportar = [
    'RESPONSABLE DE VENTA', 'SEDE', 'ALIADO COMERCIAL', 'CUENTA CONTRATO',
    'CLIENTE', 'DNI', 'TELÉFONO', 'NÚMERO TELÉFONO OPCIONAL', 'CORREO',
    'Nro. PEDIDO VENTA', 'IMPORTE (S./)', 'Nro. DE CUOTAS', 'FECHA VENTA',
    'TIPO DESPACHO', 'ESTADO', 'FECHA SOLICITUD', 'MOTIVO', 'ESTADO ANULACION',
    'PRODUCTO_1', 'CANAL_VENTA'
]

def verificar_archivos():
    """Verifica que los archivos necesarios existan"""
    print("🔍 Verificando archivos...")
    
    if not os.path.exists(ruta_fuente_principal):
        raise FileNotFoundError(f"❌ Archivo fuente principal no encontrado: {ruta_fuente_principal}")
    print("✅ Archivo fuente principal: OK")
    
    if not os.path.exists(ruta_fuente_complementaria):
        raise FileNotFoundError(f"❌ Archivo fuente complementaria no encontrado: {ruta_fuente_complementaria}")
    print("✅ Archivo fuente complementaria: OK")
    
    if not os.path.exists(ruta_destino):
        raise FileNotFoundError(f"❌ Archivo destino no encontrado: {ruta_destino}")
    print("✅ Archivo destino: OK")
    
    return True

def cargar_datos_principal():
    """Carga y procesa los datos del archivo principal"""
    print("📂 Cargando datos principales (Base Solicitudes Anulacion)...")
    
    df = pd.read_excel(ruta_fuente_principal, sheet_name='Bandeja Anulacion', dtype=str)
    print(f"📊 Registros cargados: {len(df)}")
    
    # Limpiar espacios
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x) 
                  if col.dtype == 'object' else col)
    
    # Procesar fechas
    def convertir_fecha(fecha_str):
        if pd.isna(fecha_str) or fecha_str == '':
            return ''
        fecha_str = str(fecha_str).strip()
        return fecha_str.replace('-', '/') if '-' in fecha_str else fecha_str
    
    if 'FECHA SOLICITUD' in df.columns:
        df['FECHA SOLICITUD'] = df['FECHA SOLICITUD'].apply(convertir_fecha)
        df['FECHA SOLICITUD_DATETIME'] = pd.to_datetime(df['FECHA SOLICITUD'], format='%d/%m/%Y', errors='coerce')
    
    if 'FECHA VENTA' in df.columns:
        df['FECHA VENTA'] = df['FECHA VENTA'].apply(convertir_fecha)
    
    # Filtrar por fecha
    df = df[df['FECHA SOLICITUD_DATETIME'] >= fecha_inicio]
    print(f"📅 Registros desde {fecha_inicio.strftime('%d/%m/%Y')}: {len(df)}")
    
    return df

def cargar_datos_complementarios():
    """Carga y procesa los datos complementarios"""
    print("📂 Cargando datos complementarios...")
    
    df = pd.read_excel(ruta_fuente_complementaria, dtype=str)
    print(f"📊 Registros complementarios: {len(df)}")
    
    # Limpiar espacios
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x) 
                  if col.dtype == 'object' else col)
    
    # Concatenar productos
    df = concatenar_productos(df)
    
    return df

def concatenar_productos(df):
    """Concatena múltiples columnas de productos"""
    print("🔧 Concatenando productos...")
    
    # Buscar columnas PRODUCTO_X
    columnas_producto = [col for col in df.columns if col.startswith('PRODUCTO_')]
    
    if not columnas_producto:
        print("⚠️ No se encontraron columnas de productos múltiples")
        if 'PRODUCTO' in df.columns:
            df['PRODUCTO_1'] = df['PRODUCTO']
        else:
            df['PRODUCTO_1'] = ''
        return df
    
    # Ordenar columnas numéricamente
    columnas_producto.sort(key=lambda x: int(x.split('_')[1]) if x.split('_')[1].isdigit() else 0)
    
    # Concatenar productos por fila
    productos_concatenados = []
    for _, fila in df.iterrows():
        productos = []
        for col in columnas_producto:
            valor = fila[col] if col in fila else ''
            if pd.notna(valor) and str(valor).strip() and str(valor).strip().upper() != 'NAN':
                productos.append(str(valor).strip())
        
        resultado = ' | '.join(productos) if productos else ''
        productos_concatenados.append(resultado)
    
    df['PRODUCTO_1'] = productos_concatenados
    print(f"✅ Productos concatenados: {len([p for p in productos_concatenados if p])} registros con datos")
    
    return df

def aplicar_filtros_solicitudes(df):
    """Aplica filtros específicos para solicitudes"""
    print("🔍 Aplicando filtros de solicitudes...")
    
    filtro = (
        (~df['MOTIVO'].str.upper().isin([m.upper() for m in motivos_excluir])) &
        (df['ESTADO'].str.upper() == 'PENDIENTE DE ANULACIÓN') &
        (df['ESTADO ANULACION'].str.upper().isin([e.upper() for e in estados_anulacion_validos]))
    )
    
    df_filtrado = df[filtro].copy()
    print(f"✅ Solicitudes válidas después de filtros: {len(df_filtrado)}")
    
    return df_filtrado

def buscar_columna_pedido(df, descripcion=""):
    """Busca la columna de número de pedido en el DataFrame"""
    posibles_nombres = [
        'Nro. PEDIDO VENTA', 'NRO. PEDIDO VENTA', 'PEDIDO VENTA',
        'NUMERO PEDIDO VENTA', 'NRO PEDIDO VENTA', 'N° PEDIDO VENTA',
        'Nº PEDIDO VENTA', 'No. PEDIDO VENTA', 'NO. PEDIDO VENTA'
    ]
    
    for col in df.columns:
        for posible in posibles_nombres:
            if col.upper().strip() == posible.upper():
                print(f"   🔍 Columna pedido en {descripcion}: '{col}'")
                return col
    
    return None

def combinar_datos(df_principal, df_complementarios):
    """Combina datos principales con complementarios por número de pedido"""
    print("🔗 Combinando datos por Nro. PEDIDO VENTA...")
    
    # Buscar columnas de pedido
    col_pedido_principal = buscar_columna_pedido(df_principal, "datos principales")
    col_pedido_comp = buscar_columna_pedido(df_complementarios, "datos complementarios")
    
    if not col_pedido_principal:
        raise KeyError("❌ No se encontró columna de pedido en datos principales")
    
    if not col_pedido_comp:
        raise KeyError("❌ No se encontró columna de pedido en datos complementarios")
    
    # Seleccionar columnas complementarias
    columnas_complementarias = [col_pedido_comp]
    for col_buscar in ['TELÉFONO', 'NÚMERO TELÉFONO OPCIONAL', 'CORREO', 'PRODUCTO_1', 'CANAL_VENTA']:
        for col_real in df_complementarios.columns:
            if col_real.upper().strip() == col_buscar.upper():
                columnas_complementarias.append(col_real)
                print(f"   ✅ Encontrada: '{col_real}'")
                break
        else:
            print(f"   ⚠️ No encontrada: '{col_buscar}'")
    
    df_comp_filtrado = df_complementarios[columnas_complementarias].copy()
    
    # Preparar para merge
    df_principal_merge = df_principal.rename(columns={col_pedido_principal: 'PEDIDO_MERGE'})
    df_comp_merge = df_comp_filtrado.rename(columns={col_pedido_comp: 'PEDIDO_MERGE'})
    
    # Realizar LEFT JOIN
    print(f"🔗 Realizando cruce: {len(df_principal_merge)} solicitudes con {len(df_comp_merge)} complementarios...")
    df_combinado = pd.merge(df_principal_merge, df_comp_merge, on='PEDIDO_MERGE', how='left', suffixes=('', '_comp'))
    
    # Restaurar nombre original
    df_combinado = df_combinado.rename(columns={'PEDIDO_MERGE': 'Nro. PEDIDO VENTA'})
    
    # Normalizar columnas complementarias
    mapeo_columnas = {
        'TELÉFONO': 'TELÉFONO',
        'NÚMERO TELÉFONO OPCIONAL': 'NÚMERO TELÉFONO OPCIONAL',
        'CORREO': 'CORREO',
        'PRODUCTO_1': 'PRODUCTO_1',
        'CANAL_VENTA': 'CANAL_VENTA'
    }
    
    for col_destino, col_origen in mapeo_columnas.items():
        col_encontrada = None
        for col in df_combinado.columns:
            if col.upper().strip() == col_origen.upper() or col.upper().strip() == f"{col_origen.upper()}_COMP":
                col_encontrada = col
                break
        
        if col_encontrada and col_encontrada != col_destino:
            df_combinado[col_destino] = df_combinado[col_encontrada].fillna('')
            if col_encontrada.endswith('_comp'):
                df_combinado.drop(columns=[col_encontrada], inplace=True)
        elif col_destino not in df_combinado.columns:
            df_combinado[col_destino] = ''
    
    cruces_exitosos = (df_combinado['TELÉFONO'].notna() & (df_combinado['TELÉFONO'] != '')).sum()
    print(f"✅ Cruces exitosos: {cruces_exitosos}/{len(df_combinado)}")
    
    return df_combinado

def formatear_datos(df):
    """Aplica formatos específicos"""
    print("🎨 Aplicando formatos...")
    
    # Formatear importes
    def formatear_importe(valor):
        if pd.isna(valor) or valor == '':
            return ''
        try:
            numero = float(str(valor).replace(',', ''))
            return f"{numero:.2f}"
        except:
            return str(valor)
    
    df['IMPORTE (S./)'] = df['IMPORTE (S./)'].apply(formatear_importe)
    
    # Formatear fechas
    def convertir_fecha(fecha_str):
        if pd.isna(fecha_str) or fecha_str == '':
            return ''
        fecha_str = str(fecha_str).strip()
        return fecha_str.replace('-', '/') if '-' in fecha_str else fecha_str
    
    for col_fecha in ['FECHA SOLICITUD', 'FECHA VENTA']:
        if col_fecha in df.columns:
            df[col_fecha] = df[col_fecha].apply(convertir_fecha)
    
    # Ordenar por fecha de solicitud
    if 'FECHA SOLICITUD' in df.columns:
        fecha_temp = pd.to_datetime(df['FECHA SOLICITUD'], format='%d/%m/%Y', errors='coerce')
        df = df.iloc[fecha_temp.argsort()].reset_index(drop=True)
        print("✅ Datos ordenados por FECHA SOLICITUD")
    
    return df

def obtener_registros_nuevos(df_procesados):
    """Identifica registros nuevos comparando con el archivo destino"""
    print("🔍 Identificando registros nuevos...")
    
    try:
        df_existente = pd.read_excel(ruta_destino, sheet_name='Solicitudes de Anulación')
        pedidos_existentes = set(df_existente['Nro. PEDIDO VENTA'].dropna().astype(str))
        
        print(f"📊 Registros existentes: {len(df_existente)}")
        print(f"📊 Pedidos únicos existentes: {len(pedidos_existentes)}")
        
        # Filtrar nuevos
        df_nuevos = df_procesados[
            ~df_procesados['Nro. PEDIDO VENTA'].astype(str).isin(pedidos_existentes)
        ].copy()
        
        duplicados = len(df_procesados) - len(df_nuevos)
        print(f"📈 Total procesados: {len(df_procesados)}")
        print(f"🔄 Ya existentes: {duplicados}")
        print(f"✨ Nuevos a agregar: {len(df_nuevos)}")
        
        return df_nuevos, len(df_existente)
        
    except Exception as e:
        print(f"❌ Error al identificar nuevos: {e}")
        raise

def agregar_al_archivo_destino(df_nuevos):
    """Agrega nuevos registros al archivo destino"""
    if df_nuevos.empty:
        print("ℹ️ No hay registros nuevos para agregar")
        return True
    
    print(f"💾 Agregando {len(df_nuevos)} registros...")
    
    try:
        wb = load_workbook(ruta_destino)
        ws = wb['Solicitudes de Anulación']
        
        # Encontrar última fila con datos
        ultima_fila = 1
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value:
                ultima_fila = fila
            else:
                break
        
        # Mapear columnas
        encabezados = [str(ws.cell(row=1, column=col).value).strip() 
                      for col in range(1, ws.max_column + 1)]
        mapeo_columnas = {enc: idx + 1 for idx, enc in enumerate(encabezados)}
        
        # Formato
        font_aptos = Font(name='Aptos', size=8)
        alignment_left = Alignment(horizontal='left', vertical='center')
        
        # Agregar datos
        fila_actual = ultima_fila + 1
        for _, row in df_nuevos.iterrows():
            for columna in columnas_exportar:
                if columna in mapeo_columnas:
                    col_idx = mapeo_columnas[columna]
                    valor = row[columna] if columna in row else ""
                    valor = "" if pd.isna(valor) else valor
                    
                    cell = ws.cell(row=fila_actual, column=col_idx, value=valor)
                    cell.font = font_aptos
                    cell.alignment = alignment_left
            
            fila_actual += 1
        
        wb.save(ruta_destino)
        wb.close()
        
        print(f"✅ {len(df_nuevos)} registros agregados exitosamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al agregar: {e}")
        traceback.print_exc()
        return False

def enviar_correo_resumen(exito, total_procesados=0, total_nuevos=0, error_msg=""):
    """Envía correo con resumen del proceso"""
    try:
        if not os.path.exists(ruta_destinatarios):
            print("❌ Archivo de destinatarios no encontrado")
            return
        
        df_dest = pd.read_excel(ruta_destinatarios)
        destinatarios = df_dest.iloc[:, 0].dropna().tolist()
        copia = df_dest.iloc[:, 1].dropna().tolist() if len(df_dest.columns) > 1 else []
        
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        
        mail.To = "; ".join(destinatarios)
        if copia:
            mail.CC = "; ".join(copia)
        
        fecha_proceso = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        if exito:
            asunto = f"✅ Actualización Solicitudes de Anulación - {fecha_proceso}"
            cuerpo = f"""<html><body style='font-family:Aptos, sans-serif; font-size:11pt;'>
Buenos días:<br><br>
Se completó exitosamente la actualización de solicitudes de anulación.<br><br>
<b>📊 Resumen:</b><br>
• Registros procesados: {total_procesados}<br>
• Registros nuevos agregados: {total_nuevos}<br>
• Proceso: Simplificado (sin archivos temporales)<br>
• Fecha/hora: {fecha_proceso}<br><br>
<b>✅ Validaciones aplicadas:</b><br>
• Filtrado por fecha (desde Sep 2025)<br>
• Estados de anulación válidos<br>
• Motivos de anulación filtrados<br>
• Cruce con datos complementarios<br>
• Productos concatenados<br>
• Formatos aplicados<br><br>
Saludos cordiales.<br><br>
<img src="cid:firmaimg">
</body></html>"""
        else:
            asunto = f"❌ Error en Solicitudes de Anulación - {fecha_proceso}"
            cuerpo = f"""<html><body style='font-family:Aptos, sans-serif; font-size:11pt;'>
Buenos días:<br><br>
Ocurrió un error durante la actualización:<br><br>
<b>❌ Error:</b><br>
{error_msg}<br><br>
<b>⏰ Fecha/hora:</b> {fecha_proceso}<br><br>
Por favor revisar.<br><br>
<img src="cid:firmaimg">
</body></html>"""
        
        mail.Subject = asunto
        mail.HTMLBody = cuerpo
        
        if os.path.exists(firma_path):
            attach = mail.Attachments.Add(firma_path)
            attach.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "firmaimg")
        
        mail.Display()
        print("📧 Correo preparado")
        
    except Exception as e:
        print(f"❌ Error en correo: {e}")

def main():
    """Proceso principal simplificado"""
    try:
        # 1. Verificar archivos
        verificar_archivos()
        
        # 2. Cargar datos principales
        df_principal = cargar_datos_principal()
        
        if df_principal.empty:
            print("⚠️ No hay datos principales para procesar")
            enviar_correo_resumen(True, 0, 0)
            return
        
        # 3. Aplicar filtros
        df_solicitudes = aplicar_filtros_solicitudes(df_principal)
        
        if df_solicitudes.empty:
            print("⚠️ No hay solicitudes válidas")
            enviar_correo_resumen(True, 0, 0)
            return
        
        # 4. Cargar y combinar con datos complementarios
        df_complementarios = cargar_datos_complementarios()
        df_combinado = combinar_datos(df_solicitudes, df_complementarios)
        
        # 5. Formatear datos
        df_final = formatear_datos(df_combinado)
        
        # 6. Preparar para exportar
        for col in columnas_exportar:
            if col not in df_final.columns:
                df_final[col] = ''
        
        df_exportar = df_final[columnas_exportar].copy()
        
        # 7. Identificar registros nuevos
        df_nuevos, registros_originales = obtener_registros_nuevos(df_exportar)
        
        # 8. Agregar al archivo destino
        if not df_nuevos.empty:
            exito = agregar_al_archivo_destino(df_nuevos)
            
            if exito:
                print("🎉 Proceso completado exitosamente!")
                print(f"📊 Resumen:")
                print(f"   • Registros base: {registros_originales}")
                print(f"   • Nuevos agregados: {len(df_nuevos)}")
                print(f"   • Total final: {registros_originales + len(df_nuevos)}")
                
                enviar_correo_resumen(True, len(df_exportar), len(df_nuevos))
            else:
                enviar_correo_resumen(False, error_msg="Error al agregar registros")
        else:
            print("ℹ️ No hay registros nuevos")
            enviar_correo_resumen(True, len(df_exportar), 0)
            
    except Exception as e:
        error_msg = f"Error: {str(e)}\n\nDetalle:\n{traceback.format_exc()}"
        print(f"❌ Error en proceso principal: {e}")
        print(traceback.format_exc())
        enviar_correo_resumen(False, error_msg=error_msg)

if __name__ == "__main__":
    main()
    print("🏁 Proceso finalizado.")