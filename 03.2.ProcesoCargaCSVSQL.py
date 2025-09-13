import pandas as pd
import numpy as np
import os
import pyodbc
from tqdm import tqdm
import logging
from openpyxl import load_workbook
import warnings
from sqlalchemy import create_engine

# === CONFIGURACIÓN GENERAL ===
CONFIG_PATH = r"D:\FNB\Proyectos\Exportado.xlsx"
CONFIG_SHEET = "Configuracion"
CHUNK_SIZE = 5000
ERRORS_OUT = r"D:\FNB\Proyectos\Errores_Carga_SQL.csv"
AUDITORIA_OUT = r"D:\FNB\Proyectos\Auditoria_Incremental.csv"

# === LOGGING ===
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger()

# === SUPRIMIR WARNINGS ===
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)


def leer_configuracion():
    wb = load_workbook(CONFIG_PATH, data_only=True)
    ws = wb[CONFIG_SHEET]
    cfg = {
        "servidor": ws["C11"].value,
        "base": ws["C12"].value,
        "usuario": ws["C13"].value,
        "clave": ws["C14"].value,
        "tabla": ws["C18"].value,
        "ruta_csv": ws["C22"].value,
        "archivo_csv": ws["C23"].value,
        "columna_clave": ws["C25"].value,
        "columnas_comparar": ws["C26"].value
    }
    wb.close()
    return cfg


def conectar_sql(cfg):
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={cfg['servidor']};DATABASE={cfg['base']};"
        f"UID={cfg['usuario']};PWD={cfg['clave']}"
    )
    return pyodbc.connect(conn_str)


def verificar_tabla_existe(conn, tabla):
    query = f"""
        SELECT COUNT(*)
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_NAME = '{tabla}'
    """
    with conn.cursor() as cursor:
        cursor.execute(query)
        return cursor.fetchone()[0] > 0


def obtener_estructura_sql(conn, tabla):
    query = f"""
        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = '{tabla}'
        ORDER BY ORDINAL_POSITION
    """
    return pd.read_sql(query, con=conn)


def convertir_tipos(df, estructura):
    """
    ✅ FUNCIÓN MEJORADA PARA MANEJAR CONVERSIÓN DESDE STRING BASE
    """
    for _, row in estructura.iterrows():
        col = row['COLUMN_NAME']
        tipo = row['DATA_TYPE']
        max_length = row.get('CHARACTER_MAXIMUM_LENGTH')

        if col not in df.columns:
            continue

        try:
            if tipo in ['int', 'bigint', 'smallint']:
                # ✅ CONVERSIÓN MEJORADA DE ENTEROS DESDE STRING
                logger.debug(f"Convirtiendo {col} de string a entero...")

                # Manejar valores nulos primero
                mask_notna = df[col].notna()

                if mask_notna.any():
                    # Intentar conversión numérica de valores no nulos
                    df.loc[mask_notna, col] = pd.to_numeric(df.loc[mask_notna, col], errors='coerce')

                    # Convertir decimales a enteros (ej: "123.0" → 123)
                    mask_numeric = df[col].notna()
                    if mask_numeric.any():
                        df.loc[mask_numeric, col] = df.loc[mask_numeric, col].astype(float).astype('Int64')

                # Asegurar tipo nullable integer
                df[col] = df[col].astype('Int64')

            elif tipo in ['float', 'decimal', 'numeric', 'money']:
                # ✅ CONVERSIÓN MEJORADA DE DECIMALES
                logger.debug(f"Convirtiendo {col} de string a decimal...")
                df[col] = pd.to_numeric(df[col], errors='coerce')

            elif tipo in ['date', 'datetime', 'datetime2', 'smalldatetime']:
                # ✅ CONVERSIÓN MEJORADA DE FECHAS
                logger.debug(f"Convirtiendo {col} de string a fecha...")
                df[col] = pd.to_datetime(df[col], errors='coerce')

            elif tipo in ['varchar', 'nvarchar', 'char', 'nchar', 'text', 'ntext']:
                # ✅ MANEJO MEJORADO DE STRINGS
                logger.debug(f"Procesando {col} como string/nvarchar...")

                # Truncar si excede longitud máxima
                if max_length and max_length > 0:
                    mask_notna = df[col].notna()
                    if mask_notna.any():
                        # Convertir a string y truncar
                        df.loc[mask_notna, col] = df.loc[mask_notna, col].astype(str).str[:int(max_length)]
                        logger.debug(f"   Truncando {col} a máximo {max_length} caracteres")

                # Mantener como object para preservar NaN
                df[col] = df[col].astype('object')

            elif tipo in ['bit']:
                # ✅ MANEJO DE CAMPOS BOOLEAN/BIT
                logger.debug(f"Convirtiendo {col} a bit/boolean...")
                # Convertir valores comunes de boolean
                mask_notna = df[col].notna()
                if mask_notna.any():
                    df.loc[mask_notna, col] = df.loc[mask_notna, col].astype(str).str.lower()
                    df.loc[df[col].isin(['true', '1', 'yes', 'y', 's']), col] = 1
                    df.loc[df[col].isin(['false', '0', 'no', 'n']), col] = 0
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            else:
                # ✅ OTROS TIPOS - MANTENER COMO OBJECT
                logger.debug(f"Procesando {col} como tipo genérico ({tipo})...")
                df[col] = df[col].astype('object')

        except Exception as e:
            logger.warning(f"⚠️ Error convirtiendo columna {col} (tipo: {tipo}): {e}")
            # Mantener como object en caso de error
            df[col] = df[col].astype('object')

    return df


def truncar_tabla(conn, tabla):
    with conn.cursor() as cursor:
        cursor.execute(f"TRUNCATE TABLE {tabla}")
        conn.commit()
        logger.info(f"🗑️ Tabla {tabla} truncada exitosamente")


def insertar_optimizado(df, conn, tabla):
    """
    ✅ FUNCIÓN MEJORADA PARA INSERCIÓN OPTIMIZADA
    """
    if df.empty:
        logger.warning("⚠️ DataFrame vacío, no hay datos para insertar")
        return []

    cursor = conn.cursor()
    cursor.fast_executemany = True
    columnas = df.columns.tolist()
    placeholders = ','.join(['?'] * len(columnas))
    sql = f"INSERT INTO {tabla} ({','.join(columnas)}) VALUES ({placeholders})"

    logger.info(f"🚀 Iniciando inserción de {len(df):,} registros en chunks de {CHUNK_SIZE:,}")

    errores = []
    registros_insertados = 0

    for i in tqdm(range(0, len(df), CHUNK_SIZE), desc="🚀 Insertando registros"):
        chunk = df.iloc[i:i + CHUNK_SIZE].copy()

        # ✅ PRESERVAR CORRECTAMENTE LOS VALORES NONE/NULL
        chunk_values = []
        for _, row in chunk.iterrows():
            row_values = []
            for val in row:
                if pd.isna(val):
                    row_values.append(None)
                elif isinstance(val, (pd.Timestamp, np.datetime64)):
                    # Manejar fechas correctamente
                    if pd.isna(val):
                        row_values.append(None)
                    else:
                        row_values.append(val.to_pydatetime() if hasattr(val, 'to_pydatetime') else val)
                else:
                    row_values.append(val)
            chunk_values.append(row_values)

        try:
            cursor.executemany(sql, chunk_values)
            conn.commit()
            registros_insertados += len(chunk)
            logger.debug(f"✅ Chunk {i // CHUNK_SIZE + 1}: {len(chunk)} registros insertados")

        except Exception as e:
            error_msg = str(e)
            errores.append((i, error_msg))
            logger.error(f"❌ Error en chunk {i // CHUNK_SIZE + 1} (registros {i}-{i + len(chunk) - 1}): {error_msg}")
            conn.rollback()

    logger.info(f"✅ Inserción completada: {registros_insertados:,} registros insertados exitosamente")
    if errores:
        logger.warning(f"⚠️ {len(errores)} chunks fallaron durante la inserción")

    return errores


def main():
    logger.info("=== Inicio del proceso optimizado a SQL Server ===")
    try:
        cfg = leer_configuracion()
        ruta_csv = os.path.join(cfg["ruta_csv"], cfg["archivo_csv"])

        if not os.path.exists(ruta_csv):
            logger.error(f"❌ Archivo CSV no encontrado: {ruta_csv}")
            return

        # ✅ LECTURA CORREGIDA DEL CSV - FORZAR TODO COMO STRING INICIALMENTE
        logger.info(f"📁 Leyendo CSV: {ruta_csv}")

        # Primera lectura: obtener solo los nombres de columnas
        try:
            df_sample = pd.read_csv(ruta_csv, sep='|', nrows=0)
            columnas = df_sample.columns.tolist()
            logger.info(
                f"📋 Columnas detectadas en CSV: {len(columnas)} → {columnas[:5]}{'...' if len(columnas) > 5 else ''}")
        except Exception as e:
            logger.error(f"❌ Error leyendo estructura del CSV: {e}")
            return

        # Segunda lectura: forzar TODAS las columnas como string para evitar inferencia automática
        dtype_dict = {col: 'str' for col in columnas}

        try:
            df = pd.read_csv(
                ruta_csv,
                sep='|',
                dtype=dtype_dict,  # ✅ CLAVE: Forzar todo como string
                keep_default_na=False,  # ✅ No convertir automáticamente a NaN
                na_values=[],  # ✅ Lista vacía de valores NA
            )
        except Exception as e:
            logger.error(f"❌ Error leyendo datos del CSV: {e}")
            return

        # ✅ PROCESAMIENTO POST-LECTURA MEJORADO
        logger.info("🧹 Limpiando y procesando datos...")
        for col in df.columns:
            # Reemplazar valores que representan NULL/vacío
            mask_null = df[col].isin(['', 'NULL', 'null', 'NaN', 'nan', '<NA>', 'None'])
            df.loc[mask_null, col] = np.nan

            # Limpiar espacios en blanco de strings válidos
            mask_valid = ~df[col].isna()
            if mask_valid.any():
                df.loc[mask_valid, col] = df.loc[mask_valid, col].astype(str).str.strip()
                # Después del strip, si queda vacío, convertir a NaN
                mask_empty_after_strip = (df[col] == '') & mask_valid
                df.loc[mask_empty_after_strip, col] = np.nan

        logger.info(f"📊 Datos leídos: {len(df):,} filas, {len(df.columns)} columnas")

        # ✅ MOSTRAR TIPOS DESPUÉS DE LECTURA CORREGIDA
        logger.info("🔍 Tipos después de lectura CSV (todos deberían ser 'object'):")
        for col in df.columns[:10]:  # Mostrar solo primeras 10 columnas
            logger.info(f"   • {col}: {df[col].dtype}")
        if len(df.columns) > 10:
            logger.info(f"   ... y {len(df.columns) - 10} columnas más")

        # ✅ CONEXIÓN Y VALIDACIONES SQL
        with conectar_sql(cfg) as conn:
            logger.info(f"🔗 Conectado a SQL Server: {cfg['servidor']}.{cfg['base']}")

            if verificar_tabla_existe(conn, cfg["tabla"]):
                logger.info(f"✅ La tabla '{cfg['tabla']}' existe en la base de datos.")
            else:
                logger.error(f"❌ La tabla '{cfg['tabla']}' NO existe en la base de datos.")
                return

            estructura_sql = obtener_estructura_sql(conn, cfg["tabla"])
            logger.info(f"📋 Estructura SQL obtenida: {len(estructura_sql)} columnas")

            # ✅ MOSTRAR MAPEO DE TIPOS ANTES DE CONVERSIÓN
            logger.info("🔍 Mapeo de tipos SQL Server vs Pandas ANTES de conversión:")
            for _, row in estructura_sql.iterrows():
                col = row['COLUMN_NAME']
                tipo_sql = row['DATA_TYPE']
                if col in df.columns:
                    tipo_pandas = df[col].dtype
                    logger.info(f"   • {col}: {tipo_sql} ← {tipo_pandas}")
                else:
                    logger.warning(f"   ⚠️ {col}: {tipo_sql} ← COLUMNA NO ENCONTRADA EN CSV")

            # ✅ APLICAR CONVERSIÓN DE TIPOS
            logger.info("🔄 Aplicando conversiones de tipos...")
            df = convertir_tipos(df, estructura_sql)

            # ✅ MOSTRAR TIPOS DESPUÉS DE CONVERSIÓN
            logger.info("🔍 Tipos DESPUÉS de conversión:")
            for _, row in estructura_sql.iterrows():
                col = row['COLUMN_NAME']
                tipo_sql = row['DATA_TYPE']
                if col in df.columns:
                    tipo_pandas = df[col].dtype
                    # Validar que el tipo sea correcto
                    tipo_ok = "✅" if not (
                                tipo_sql in ['varchar', 'nvarchar'] and 'float' in str(tipo_pandas).lower()) else "❌"
                    logger.info(f"   {tipo_ok} {col}: {tipo_sql} → {tipo_pandas}")

            # ✅ FILTRAR SOLO COLUMNAS QUE EXISTEN EN LA TABLA SQL
            columnas_sql = estructura_sql['COLUMN_NAME'].tolist()
            columnas_faltantes = [col for col in columnas_sql if col not in df.columns]
            columnas_extra = [col for col in df.columns if col not in columnas_sql]

            if columnas_faltantes:
                logger.warning(f"⚠️ Columnas en SQL pero no en CSV: {columnas_faltantes}")
            if columnas_extra:
                logger.warning(f"⚠️ Columnas en CSV pero no en SQL: {columnas_extra}")
                # Filtrar columnas extra
                df = df[[col for col in df.columns if col in columnas_sql]]
                logger.info(f"🔧 DataFrame filtrado a {len(df.columns)} columnas válidas")

            # ✅ SELECCIÓN DE MODO
            print("\n" + "=" * 50)
            print("SELECCIÓN DE MODO DE CARGA")
            print("=" * 50)
            print("R - Reemplazar: Borra todos los datos y carga desde cero")
            print("A - Agregar: Añade datos sin validaciones (riesgo de duplicados)")
            print("I - Incremental: Solo inserta/actualiza registros nuevos o modificados")
            print("C - Cancelar: Termina sin hacer cambios")
            print("=" * 50)

            modo = input("Selecciona modo (R/A/I/C): ").strip().upper()
            if modo not in ['R', 'A', 'I', 'C']:
                logger.error("❌ Opción no válida.")
                return
            if modo == 'C':
                logger.info("🚫 Operación cancelada por el usuario.")
                return

            # ✅ PROCESAMIENTO SEGÚN MODO SELECCIONADO
            if modo == 'R':
                logger.info(f"🔁 MODO REEMPLAZAR: Eliminando contenido de '{cfg['tabla']}'...")
                truncar_tabla(conn, cfg["tabla"])
                errores = insertar_optimizado(df, conn, cfg["tabla"])

            elif modo == 'A':
                logger.info(f"➕ MODO AGREGAR: Añadiendo datos a la tabla '{cfg['tabla']}'...")
                errores = insertar_optimizado(df, conn, cfg["tabla"])

            elif modo == 'I':
                logger.info(f"🔄 MODO INCREMENTAL: Procesando cambios en '{cfg['tabla']}'...")

                clave = cfg["columna_clave"]
                columnas_actualizar = cfg["columnas_comparar"]

                if not clave or not columnas_actualizar:
                    logger.warning("⚠️ Columna clave o columnas de comparación no especificadas.")
                    print("\nPara modo incremental necesitas configurar:")
                    print("- Columna clave (C25 en Excel)")
                    print("- Columnas a comparar (C26 en Excel, separadas por ';')")

                    modo_alt = input("¿Deseas (C)ancelar, (R)eemplazar o (A)gregar?: ").strip().upper()
                    if modo_alt == 'R':
                        truncar_tabla(conn, cfg["tabla"])
                        errores = insertar_optimizado(df, conn, cfg["tabla"])
                    elif modo_alt == 'A':
                        errores = insertar_optimizado(df, conn, cfg["tabla"])
                    else:
                        logger.info("🚫 Operación cancelada.")
                        return
                else:
                    clave = clave.strip()
                    columnas = [c.strip() for c in columnas_actualizar.split(';')]

                    logger.info(f"🔑 Columna clave: {clave}")
                    logger.info(f"🔍 Columnas a comparar: {columnas}")

                    # Verificar que la columna clave existe
                    if clave not in df.columns:
                        logger.error(f"❌ Columna clave '{clave}' no encontrada en CSV")
                        return

                    # Obtener datos actuales de SQL
                    try:
                        query = f"SELECT {clave}, {','.join(columnas)} FROM {cfg['tabla']}"
                        df_sql = pd.read_sql(query, conn)
                        logger.info(f"📊 Datos actuales en SQL: {len(df_sql):,} registros")
                    except Exception as e:
                        logger.error(f"❌ Error consultando datos SQL: {e}")
                        return

                    # Merge para identificar cambios
                    df_merge = df.merge(df_sql, on=clave, how='left', suffixes=('', '_sql'))

                    # Identificar registros nuevos
                    df_nuevos = df_merge[df_merge[[col + '_sql' for col in columnas]].isnull().all(axis=1)]

                    # Identificar registros modificados
                    df_cambios = df_merge[
                        (~df_merge[[col + '_sql' for col in columnas]].isnull().all(axis=1)) &
                        (df_merge[[col for col in columnas]].astype(str).values !=
                         df_merge[[col + '_sql' for col in columnas]].astype(str).values).any(axis=1)
                        ]

                    logger.info(f"📈 Análisis incremental:")
                    logger.info(f"   • Registros nuevos: {len(df_nuevos):,}")
                    logger.info(f"   • Registros modificados: {len(df_cambios):,}")
                    logger.info(f"   • Total a procesar: {len(df_nuevos) + len(df_cambios):,}")

                    # Generar auditoría
                    auditoria = []

                    # Auditoría de cambios
                    for _, row in df_cambios.iterrows():
                        for col in columnas:
                            old = row.get(col + '_sql')
                            new = row.get(col)
                            if pd.notna(old) and str(old) != str(new):
                                auditoria.append({
                                    'Tipo': 'ACTUALIZADO',
                                    'Clave': row[clave],
                                    'Columna_Modificada': col,
                                    'Valor_Anterior': old,
                                    'Valor_Nuevo': new
                                })

                    # Auditoría de inserciones
                    for _, row in df_nuevos.iterrows():
                        auditoria.append({
                            'Tipo': 'INSERTADO',
                            'Clave': row[clave],
                            'Columna_Modificada': '',
                            'Valor_Anterior': '',
                            'Valor_Nuevo': ''
                        })

                    # Guardar auditoría
                    if auditoria:
                        df_auditoria = pd.DataFrame(auditoria)
                        df_auditoria.to_csv(AUDITORIA_OUT, index=False, encoding='utf-8-sig')
                        logger.info(f"🕵️‍♂️ Auditoría generada: {AUDITORIA_OUT} ({len(auditoria)} cambios)")

                    # Preparar datos para inserción
                    df_insert = pd.concat([df_nuevos, df_cambios])
                    if not df_insert.empty:
                        # Filtrar solo columnas de la tabla SQL
                        df_insert = df_insert[[col for col in df_insert.columns if col in columnas_sql]]
                        errores = insertar_optimizado(df_insert, conn, cfg["tabla"])
                    else:
                        logger.info("ℹ️ No hay cambios que procesar")
                        errores = []

        # ✅ MANEJO FINAL DE ERRORES
        if errores:
            logger.error(f"❌ Proceso completado con errores: {len(errores)} chunks fallidos")
            df_errores = pd.DataFrame(errores, columns=["Chunk_Inicio", "Error"])
            df_errores.to_csv(ERRORS_OUT, index=False, encoding='utf-8-sig')
            logger.info(f"📄 Detalle de errores guardado en: {ERRORS_OUT}")
        else:
            logger.info("✅ Proceso completado exitosamente sin errores.")

        logger.info("=== Fin del proceso ===")

    except Exception as e:
        logger.exception(f"❌ Error crítico en el proceso principal: {e}")


if __name__ == "__main__":
    main()