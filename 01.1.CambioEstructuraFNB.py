# procesamiento_datos_ventas.py

import pandas as pd
import numpy as np
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import warnings
from typing import Dict, List
import logging
import time
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

warnings.filterwarnings('ignore')


class SalesDataProcessor:
    def __init__(self):
        self.columnas_grupo2_fijas = [
            "PRODUCTO", "SKU", "CANTIDAD", "PRECIO", "CATEGORIA", "MARCA", "SUBCANAL",
            "CATEGORIA REAL", "TIPO PRODUCTO", "MODELO PRODUCTO", "SKU2", "DESCRIPCION"
        ]
        self.columnas_numericas = ["IMPORTE (S./)", "CRÉDITO UTILIZADO", "Nro. DE CUOTAS", "CANTIDAD", "PRECIO"]
        self.ruta_canal_fija = r"D:\\FNB\\Reportes\\19. Reportes IBR\\00. Estructura Reporte\\Canal\\Canal.xlsx"
        self.ruta_base = r"D:\\FNB\\Reportes\\19. Reportes IBR\\00. Estructura Reporte\\Base\\Base_Origen.xlsx"
        self.ruta_salida = r"D:\\FNB\\Reportes\\19. Reportes IBR\\00. Estructura Reporte\\Procesado"
        self.rangos_hora = [f"{h:02d}:{m:02d} - {(h + (m + 30) // 60) % 24:02d}:{(m + 30) % 60:02d}"
                            for h in range(24) for m in range(0, 60, 30)]
        self._cache_rango_hora: Dict[str, str] = {}
        self.columnas_grupo1 = []
        self.columnas_grupo2 = []
        self.max_productos = 4
        self.ruta_feriados = r"D:\FNB\Reportes\19. Reportes IBR\01. Pendientes de Entrega\Feriados"

    def _seleccionar_archivo(self, titulo: str) -> str:
        # Verificar si existe el archivo en la ruta predeterminada
        if os.path.exists(self.ruta_base):
            logger.info(f"Usando archivo predeterminado: {self.ruta_base}")
            return self.ruta_base
        else:
            logger.warning(f"Archivo predeterminado no encontrado: {self.ruta_base}")
            root = tk.Tk()
            root.withdraw()
            return filedialog.askopenfilename(
                title=titulo,
                filetypes=[("Archivos Excel", "*.xlsx")],
                initialdir=os.path.dirname(self.ruta_base)
            )

    def _determinar_columnas_dinamicas(self, columnas_archivo: List[str]):
        self.columnas_grupo2 = [col for col in self.columnas_grupo2_fijas if col in columnas_archivo]
        self.columnas_grupo1 = [col for col in columnas_archivo if col not in self.columnas_grupo2]
        logger.info(f"Columnas Grupo 1 (datos principales): {len(self.columnas_grupo1)}")
        logger.info(f"Columnas Grupo 2 (productos): {len(self.columnas_grupo2)}")

    def _optimizar_tipos_datos(self, columnas_disponibles: List[str]) -> Dict[str, str]:
        dtypes_dict = {}
        columnas_fecha = [col for col in columnas_disponibles if 'FECHA' in col.upper()]
        columnas_hora = [col for col in columnas_disponibles if 'HORA' in col.upper()]
        for col in columnas_disponibles:
            if col not in self.columnas_numericas and col not in columnas_fecha and col not in columnas_hora:
                dtypes_dict[col] = 'string'
        for col in self.columnas_numericas:
            if col in columnas_disponibles:
                dtypes_dict[col] = 'float32'
        return dtypes_dict

    def _actualizar_archivo_canal(self):
        try:
            if not os.path.exists(self.ruta_canal_fija):
                logger.warning(f"Archivo Canal no encontrado en: {self.ruta_canal_fija}")
                return

            logger.info("Actualizando archivo Canal...")

            # Usar una nueva instancia de Excel sin interferir con otras
            app = xw.App(visible=False, add_book=False)

            try:
                # Abrir específicamente nuestro archivo
                wb = app.books.open(self.ruta_canal_fija)

                # Actualizar solo este workbook
                wb.api.RefreshAll()
                time.sleep(3)

                # Guardar y cerrar solo nuestro archivo
                wb.save()
                wb.close()

                logger.info("Archivo Canal actualizado correctamente")

            except Exception as e:
                logger.error(f"Error procesando Canal.xlsx: {e}")
                # Intentar cerrar el workbook si está abierto
                try:
                    if 'wb' in locals():
                        wb.close()
                except:
                    pass
            finally:
                # Cerrar solo nuestra instancia de Excel
                try:
                    app.quit()
                except:
                    pass

        except Exception as e:
            logger.error(f"Error iniciando Excel para Canal: {e}")
            logger.info("Continuando sin actualizar Canal...")

    def _generar_codigo_unico_vectorizado(self, df: pd.DataFrame) -> pd.Series:
        return pd.util.hash_pandas_object(df[self.columnas_grupo1].astype(str).fillna(''), index=False)

    def _crear_mapeo_codigos(self, codigos_unicos: pd.Series) -> Dict[int, str]:
        unique_codes = codigos_unicos.unique()
        return {codigo: f"C{i + 1:07d}" for i, codigo in enumerate(unique_codes)}

    def _asignar_rango_hora_vectorizado(self, series_hora: pd.Series) -> pd.Series:
        def get_rango_hora(hora):
            if pd.isna(hora):
                return ""
            hora_str = str(hora)
            if hora_str in self._cache_rango_hora:
                return self._cache_rango_hora[hora_str]
            try:
                hora_obj = pd.to_datetime(hora).time()
                minutos = hora_obj.hour * 60 + hora_obj.minute
                idx = min(minutos // 30, len(self.rangos_hora) - 1)
                resultado = self.rangos_hora[idx]
                self._cache_rango_hora[hora_str] = resultado
                return resultado
            except:
                return ""

        return series_hora.apply(get_rango_hora)

    def _cargar_mapeo_canales(self) -> Dict[str, str]:
        try:
            df_canal = pd.read_excel(self.ruta_canal_fija, sheet_name='Hoja1')
            if len(df_canal.columns) >= 3:
                # Crear mapeo case-insensitive
                mapeo = {}
                for _, row in df_canal.iterrows():
                    clave_original = str(row.iloc[0]).strip()
                    clave_upper = clave_original.upper()
                    valor = str(row.iloc[2]).strip()
                    mapeo[clave_upper] = valor
                return mapeo
        except Exception as e:
            logger.warning(f"Error leyendo Canal.xlsx: {e}")
        return {}

    def _cargar_feriados(self) -> set:
        try:
            archivo = os.path.join(self.ruta_feriados, "Feriados.xlsx")
            if not os.path.exists(archivo):
                logger.warning(f"No se encontró archivo de feriados: {archivo}")
                return set()
            df = pd.read_excel(archivo)
            fechas = pd.to_datetime(df.iloc[:, 0], errors='coerce').dt.date
            feriados = set(f for f in fechas if not pd.isna(f))
            logger.info(f"Feriados cargados: {len(feriados)}")
            return feriados
        except Exception as e:
            logger.error(f"Error cargando feriados: {e}")
            return set()

    def _calcular_dias_habiles(self, fecha_venta: pd.Timestamp, feriados: set) -> int:
        if pd.isna(fecha_venta):
            return 0
        hoy = datetime.now().date()
        if isinstance(fecha_venta, pd.Timestamp):
            fecha_venta = fecha_venta.date()
        if fecha_venta > hoy:
            return 0
        dias_totales = pd.date_range(fecha_venta, hoy, freq='D')
        dias_habiles = [d for d in dias_totales if d.weekday() < 6 and d.date() not in feriados]
        return max(len(dias_habiles) - 1, 0)

    def _calcular_tiempo_exacto(self, fecha_venta, fecha_entrega, feriados):
        hoy = datetime.now().date()

        if pd.isna(fecha_venta):
            return 0

        inicio = fecha_venta.date() if isinstance(fecha_venta, pd.Timestamp) else fecha_venta
        if pd.isna(inicio):
            return 0

        if pd.isna(fecha_entrega):
            fin = hoy
        else:
            fin = fecha_entrega.date() if isinstance(fecha_entrega, pd.Timestamp) else fecha_entrega

        if inicio > fin:
            return 0

        dias_totales = pd.date_range(inicio, fin, freq='D')
        # Solo se excluyen domingos (weekday=6) y feriados
        dias_habiles = [d for d in dias_totales if d.weekday() != 6 and d.date() not in feriados]

        resultado = len(dias_habiles) - 1
        return 0 if resultado == -1 else resultado

    def _determinar_canal_venta_vectorizado(self, df: pd.DataFrame, mapeo: Dict[str, str]) -> tuple:
        """
        Determina el canal de venta y retorna también las sedes sin canal asignado
        """
        responsable = df['RESPONSABLE DE VENTA'].astype(str).str.strip().str.upper()
        aliado = df['ALIADO COMERCIAL'].astype(str).str.strip().str.upper()
        fecha_venta = pd.to_datetime(df['FECHA VENTA'], errors='coerce')
        sede = df['SEDE'].astype(str).str.strip().str.upper()
        categoria = df.get('CATEGORIA_1', pd.Series([''] * len(df))).astype(str).str.strip().str.upper()

        canal = pd.Series([''] * len(df))
        fecha_limite = pd.to_datetime('2024-02-01')
        fecha_limite_1 = pd.to_datetime('2025-08-01')

        # Condiciones específicas
        cond_retail_1 = (fecha_venta >= fecha_limite_1) & (responsable.isin(["TOPITOP"]))
        cond_retail = (fecha_venta >= fecha_limite) & (responsable.isin(["CONECTA RETAIL", "INTEGRA RETAIL"]))
        cond_materiales = (categoria == "MATERIALES Y ACABADOS DE CONSTRUCCIÓN") & (~responsable.isin(
            ["A & G INGENIERIA", "INCOSER GAS PERU S.A.C.", "PROMART"]))
        cond_motos = (
            categoria.isin(["MOTOS", "MOTOS ELECTRICAS", "ACCESORIOS MOTOS"]) &
            (~responsable.isin(["CONECTA RETAIL", "INTEGRA RETAIL"]))
        )
        cond_merpes = (aliado == "GRUPO MERPES") & (categoria == "MUEBLES")

        # Asignar canales
        canal.loc[cond_retail] = "RETAIL"
        canal.loc[cond_retail_1] = "RETAIL"
        canal.loc[cond_materiales] = "MATERIALES Y ACABADOS DE CONSTRUCCIÓN"
        canal.loc[cond_motos] = "MOTOS"
        canal.loc[cond_merpes] = "MATERIALES Y ACABADOS DE CONSTRUCCIÓN"

        # Para los que no tienen canal asignado, usar mapeo de SEDE
        mask_sin_canal = canal == ''
        canal_desde_sede = sede.loc[mask_sin_canal].map(mapeo).fillna('')
        canal.loc[mask_sin_canal] = canal_desde_sede

        # Identificar sedes que no tienen canal asignado
        mask_final_sin_canal = canal == ''
        sedes_sin_canal = set()
        if mask_final_sin_canal.any():
            sedes_sin_canal = set(sede.loc[mask_final_sin_canal].unique())
            # Remover valores vacíos o nan
            sedes_sin_canal = {s for s in sedes_sin_canal if s and str(s).strip() and str(s).upper() != 'NAN'}

        return canal, sedes_sin_canal

    def _crear_columnas_pivotadas_ordenadas(self, df_g2_pivot: pd.DataFrame) -> pd.DataFrame:
        """Reordena las columnas pivotadas para que estén agrupadas por producto"""
        columnas_base = ['codigo_numerado']

        # Obtener todas las columnas pivotadas
        columnas_pivotadas = [col for col in df_g2_pivot.columns if col != 'codigo_numerado']

        # Crear lista ordenada de columnas finales
        columnas_finales = columnas_base.copy()

        # Agrupar por número de producto
        for i in range(1, self.max_productos + 1):
            for col_base in self.columnas_grupo2_fijas:
                col_nombre = f"{col_base}_{i}"
                if col_nombre in columnas_pivotadas:
                    columnas_finales.append(col_nombre)

        # Asegurar que todas las columnas estén presentes
        for col in columnas_pivotadas:
            if col not in columnas_finales:
                columnas_finales.append(col)

        return df_g2_pivot[columnas_finales]

    def _parsear_fecha_dd_mm_yyyy(self, fecha_str):
        """
        Parsea fechas en formato dd/mm/yyyy específicamente
        """
        if pd.isna(fecha_str) or fecha_str == '':
            return pd.NaT

        try:
            fecha_str = str(fecha_str).strip()
            # Si ya es datetime, convertir a string primero
            if isinstance(fecha_str, pd.Timestamp):
                return fecha_str

            # Intentar varios formatos de fecha
            formatos = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d']

            for formato in formatos:
                try:
                    return pd.to_datetime(fecha_str, format=formato)
                except:
                    continue

            # Si no funciona ningún formato específico, usar dayfirst=True
            return pd.to_datetime(fecha_str, dayfirst=True, errors='coerce')

        except Exception as e:
            logger.warning(f"Error parseando fecha '{fecha_str}': {e}")
            return pd.NaT

    def _aplicar_formato_excel(self, archivo_salida: str):
        """Aplica formato específico al archivo Excel"""
        try:
            wb = load_workbook(archivo_salida)
            ws = wb.active

            # Aplicar formato de fecha a columnas de fecha
            from openpyxl.styles import NamedStyle
            date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and 'FECHA' in str(col_name.value).upper():
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].style = date_style
                        except:
                            pass

            # Aplicar formato de hora a columnas de hora (formato texto para mostrar solo HH:MM)
            text_time_style = NamedStyle(name='text_time_style', number_format='@')

            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and 'HORA' in str(col_name.value).upper():
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].style = text_time_style
                        except:
                            pass

            # Aplicar formato numérico a columnas de importes
            currency_style = NamedStyle(name='currency_style', number_format='#,##0.00')

            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and col_name.value in ["IMPORTE (S./)", "CRÉDITO UTILIZADO"]:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].style = currency_style
                        except:
                            pass

            # Definir fuente y aplicar a todas las celdas
            font_aptos = Font(name='Aptos', size=8)

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font_aptos

            # Ajustar altura de filas
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 11.25

            # Autoajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                # Ajustar ancho con un mínimo de 8 y máximo de 50
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(archivo_salida)
            logger.info("Formato aplicado correctamente al archivo Excel")

        except Exception as e:
            logger.error(f"Error aplicando formato Excel: {e}")

    def procesar(self):
        try:
            archivo = self._seleccionar_archivo("Seleccione archivo de ventas")
            if not archivo:
                logger.error("No se seleccionó archivo")
                return

            logger.info(f"Archivo seleccionado: {archivo}")
            self._actualizar_archivo_canal()

            columnas = pd.read_excel(archivo, nrows=0).columns.tolist()
            self._determinar_columnas_dinamicas(columnas)

            if not self.columnas_grupo2:
                raise ValueError("No se encontraron columnas de productos")

            dtypes = self._optimizar_tipos_datos(columnas)
            columnas_fecha = [col for col in columnas if 'FECHA' in col.upper()]
            columnas_hora = [col for col in columnas if 'HORA' in col.upper()]

            # CORRECCIÓN: Cargar fechas como string para parsear manualmente
            df = pd.read_excel(archivo, dtype=dtypes)
            if df.empty:
                raise ValueError("Archivo sin datos")

            logger.info(f"Datos cargados: {len(df):,} filas")

            # CORRECCIÓN: Parsear fechas manualmente con formato dd/mm/yyyy
            for fecha_col in columnas_fecha:
                if fecha_col in df.columns:
                    logger.info(f"Parseando columna de fecha: {fecha_col}")
                    df[fecha_col] = df[fecha_col].apply(self._parsear_fecha_dd_mm_yyyy)
                    logger.info(f"Fechas parseadas para {fecha_col}")

            # Parsear columnas de hora
            for hora_col in columnas_hora:
                if hora_col in df.columns:
                    df[hora_col] = pd.to_datetime(df[hora_col], errors='coerce')

            # Generar códigos únicos
            df['codigo_unico'] = self._generar_codigo_unico_vectorizado(df)
            dict_codigos = self._crear_mapeo_codigos(df['codigo_unico'])
            df['codigo_numerado'] = df['codigo_unico'].map(dict_codigos)

            # Crear dataframes separados
            df_g1 = df.drop_duplicates('codigo_unico')[['codigo_numerado'] + self.columnas_grupo1].copy()
            df_g2 = df[['codigo_unico'] + self.columnas_grupo2].copy()
            df_g2['codigo_numerado'] = df_g2['codigo_unico'].map(dict_codigos)
            df_g2['producto_idx'] = df_g2.groupby('codigo_unico').cumcount() + 1
            df_g2 = df_g2[df_g2['producto_idx'] <= self.max_productos]

            # Pivotar datos de productos
            df_g2_long = pd.melt(df_g2, id_vars=['codigo_numerado', 'producto_idx'],
                                 value_vars=self.columnas_grupo2, var_name='atributo', value_name='valor')
            df_g2_long['col_final'] = df_g2_long['atributo'] + '_' + df_g2_long['producto_idx'].astype(str)
            df_g2_pivot = df_g2_long.pivot(index='codigo_numerado', columns='col_final', values='valor').reset_index()

            # Reordenar columnas pivotadas
            df_g2_pivot = self._crear_columnas_pivotadas_ordenadas(df_g2_pivot)

            # Unir dataframes
            df_final = df_g1.merge(df_g2_pivot, on='codigo_numerado', how='left')

            # Agregar rango de hora si existe
            if 'HORA VENTA' in df_final.columns:
                df_final['RANGO HORA'] = self._asignar_rango_hora_vectorizado(df_final['HORA VENTA'])

            # === CÁLCULO DE TIEMPO Y RANGO DE ENTREGA ===
            feriados = self._cargar_feriados()

            # Cálculo de TipoProducto
            df_final['TipoProducto'] = df_final['PRODUCTO_1'].astype(str).str.upper().str.contains(
                "PUNTO|DUCTE|ADICIONAL", na=False)
            df_final['TipoProducto'] = np.where(df_final['TipoProducto'], "CON CONSTRUCCIÓN", "PRODUCTO SOLO")

            # Cálculo de Tipo Producto
            df_final['Tipo Producto'] = np.where(
                (df_final['ALIADO COMERCIAL'].astype(str).str.upper() == "GASODOMESTICOS") &
                (df_final['TipoProducto'] == "CON CONSTRUCCIÓN"),
                "CON CONSTRUCCIÓN", "PRODUCTO SOLO"
            )

            # Cálculo de tiempo en días hábiles (de FECHA VENTA a FECHA ENTREGA o HOY)
            col_fecha_entrega = 'FECHA ENTREGA' if 'FECHA ENTREGA' in df_final.columns else None

            if col_fecha_entrega:
                df_final['Tiempo'] = df_final.apply(
                    lambda row: self._calcular_tiempo_exacto(row['FECHA VENTA'], row[col_fecha_entrega], feriados),
                    axis=1
                )
            else:
                df_final['Tiempo'] = df_final.apply(
                    lambda row: self._calcular_tiempo_exacto(row['FECHA VENTA'], pd.NaT, feriados),
                    axis=1
                )

            # Evaluación de rango de entrega basado en 'Tiempo'
            def evaluar_rango(row):
                aliado = str(row.get('ALIADO COMERCIAL', '')).upper()
                canal = str(row.get('CANAL_VENTA', '')).upper()
                categoria = str(row.get('CATEGORIA_1', '')).upper()
                tipo_producto = row.get('Tipo Producto', '')
                tiempo = row.get('Tiempo', 0)

                if pd.isna(tiempo):
                    return ""

                if aliado == 'MALL HOGAR':
                    return "FUERA DE PLAZO" if tiempo > 10 else "DENTRO DE PLAZO"
                if canal == 'MOTOS':
                    return "FUERA DE PLAZO" if tiempo > 30 else "DENTRO DE PLAZO"
                if categoria == 'MUEBLES':
                    return "FUERA DE PLAZO" if tiempo > 15 else "DENTRO DE PLAZO"
                if tipo_producto == 'CON CONSTRUCCIÓN':
                    return "FUERA DE PLAZO" if tiempo > 15 else "DENTRO DE PLAZO"
                if tipo_producto == 'PRODUCTO SOLO':
                    return "FUERA DE PLAZO" if tiempo > 4 else "DENTRO DE PLAZO"

                return ""

            df_final['Rango'] = df_final.apply(evaluar_rango, axis=1)
            logger.info("Columnas 'TipoProducto', 'Tipo Producto', 'Tiempo' y 'Rango' agregadas exitosamente")

            # CORRECCIÓN: Determinar canal de venta y validar completitud
            mapeo = self._cargar_mapeo_canales()
            canal_result, sedes_sin_canal = self._determinar_canal_venta_vectorizado(df_final, mapeo)
            df_final['CANAL_VENTA'] = canal_result

            # Verificar que no queden canales en blanco
            if sedes_sin_canal:
                print("❌ ERROR: Se encontraron sedes sin canal asignado:")
                print("=" * 60)
                for sede in sorted(sedes_sin_canal):
                    print(f"   • {sede}")
                print("=" * 60)
                print("Por favor, actualice el archivo Canal.xlsx con estas sedes antes de continuar.")
                print("El proceso se ha detenido.")
                return

            # Verificar si aún hay registros sin canal
            registros_sin_canal = (df_final['CANAL_VENTA'] == '') | (df_final['CANAL_VENTA'].isna())
            if registros_sin_canal.any():
                count_sin_canal = registros_sin_canal.sum()
                print(f"❌ ERROR: {count_sin_canal} registros sin canal de venta asignado.")
                print("Revise la configuración del mapeo de canales.")
                return

           # === AJUSTE CARDIF (vectorizado) ===
            if 'Nro. PEDIDO CARDIF' in df_final.columns and 'Nro. PEDIDO VENTA' in df_final.columns:
                # Crear un diccionario {pedido_cardif: canal}
                mapeo_cardif = df_final.dropna(subset=['Nro. PEDIDO CARDIF']).set_index('Nro. PEDIDO CARDIF')['CANAL_VENTA'].to_dict()

                # Reemplazar en CANAL_VENTA para los pedidos que aparezcan en Nro. PEDIDO VENTA
                df_final['CANAL_VENTA'] = df_final.apply(
                    lambda row: mapeo_cardif.get(row['Nro. PEDIDO VENTA'], row['CANAL_VENTA']),
                    axis=1
                )

                logger.info("Ajuste CANAL_VENTA con Nro. PEDIDO CARDIF aplicado correctamente")

            # AJUSTE 1: Cambiar CHATBOT por DIGITAL en la columna CANAL_VENTA
            if 'CANAL_VENTA' in df_final.columns:
                df_final['CANAL_VENTA'] = df_final['CANAL_VENTA'].replace('CHATBOT', 'DIGITAL')
                logger.info("Valores 'CHATBOT' cambiados por 'DIGITAL' en CANAL_VENTA")

            # AJUSTE 2: Formatear HORA VENTA correctamente para Excel (solo hora, sin fecha)
            if 'HORA VENTA' in df_final.columns:
                # Convertir a datetime para extraer la hora
                hora_temp = pd.to_datetime(df_final['HORA VENTA'], errors='coerce')
                # Extraer solo la hora como string en formato HH:MM
                df_final['HORA VENTA'] = hora_temp.dt.strftime('%H:%M')
                # Reemplazar valores NaT con cadena vacía
                df_final['HORA VENTA'] = df_final['HORA VENTA'].fillna('')
                logger.info("Formato de HORA VENTA corregido para Excel (solo HH:MM)")

            # Formatear números con exactamente 2 decimales
            for col in ["IMPORTE (S./)", "CRÉDITO UTILIZADO"]:
                if col in df_final.columns:
                    df_final[col] = pd.to_numeric(df_final[col], errors='coerce').round(2)

            # Guardar archivo con nombre específico
            salida = os.path.join(self.ruta_salida, "Archivo_Procesado.xlsx")

            # Crear directorio si no existe
            os.makedirs(self.ruta_salida, exist_ok=True)

            # Guardar sin formato primero
            df_final.to_excel(salida, index=False)

            # Aplicar formato específico
            self._aplicar_formato_excel(salida)

            logger.info(f"Archivo generado: {salida}")
            print(f"✅ Registros procesados: {len(df_final):,}")
            print(f"✅ Productos pivotados por registro: {self.max_productos}")
            print(f"✅ CHATBOT cambiado por DIGITAL en CANAL_VENTA")
            print(f"✅ Formato de fechas corregido (dd/mm/yyyy)")
            print(f"✅ Formato de HORA VENTA corregido (solo HH:MM, sin fecha)")
            print(f"✅ Todos los registros tienen canal de venta asignado")
            print(f"✅ Formato aplicado: Fuente Aptos, tamaño 8, altura 11.25")

        except Exception as e:
            logger.error(f"Error en procesamiento: {e}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    processor = SalesDataProcessor()
    try:
        processor.max_productos = int(input("Ingrese cantidad máxima de productos por registro (1-20): "))
        if not (1 <= processor.max_productos <= 20):
            raise ValueError("Valor fuera de rango")
    except Exception:
        print("Se usará valor por defecto: 4")
        processor.max_productos = 4

    processor.procesar()