# RestructuraFNB.py
# Script para reestructurar datos de ventas con columnas específicas y validación RENIEC
# Sigue la misma lógica del 01.1 pero con nueva estructura de columnas

import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import warnings
from typing import Dict, List
import logging
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

warnings.filterwarnings('ignore')


class SalesDataRestructurer:
    def __init__(self):
        # Columnas de productos que siempre serán 4
        self.max_productos = 4
        
        # Columnas de grupo 2 (productos) que se pivotarán - DEBE INCLUIR TODAS LAS COLUMNAS DE PRODUCTOS
        self.columnas_grupo2_fijas = [
            "PRODUCTO", "SKU", "CANTIDAD", "PRECIO", "CATEGORIA", "MARCA", "SUBCANAL",
            "CATEGORIA REAL", "TIPO PRODUCTO", "MODELO PRODUCTO", "SKU2", "DESCRIPCION"
        ]
        
        # Definir el orden exacto de columnas del archivo final
        self.columnas_finales_orden = [
            "FECHA VENTA",
            "FECHA ENTREGA",
            "Columna libre 1",
            "CUENTA CONTRATO",
            "DNI",
            "CLIENTE",
            "TELÉFONO",
            "Columna libre 2",
            "CORREO",
            "Columna libre 3",
            "Columna libre 4",
            "Nro. DE CONTRATO",
            "BOLETA",
            "Columna libre 5",
            "Nro. PEDIDO VENTA",
            "Nro. PEDIDO CARDIF",
            "IMPORTE (S./)",
            "CRÉDITO UTILIZADO",
            "Nro. DE CUOTAS",
            "RESPONSABLE DE VENTA",
            "Columna libre 6",
            "ALIADO COMERCIAL",
            "Columna libre 7",
            "SEDE",
            "Columna libre 8",
            "TIPO DESPACHO",
            "ESTADO",
            "Columna libre 9",
            "Columna libre 10",
            "Columna libre 11",
            "Columna libre 12",
            "PRODUCTO_1",
            "SKU_1",
            "PRODUCTO_2",
            "SKU_2",
            "PRODUCTO_3",
            "SKU_3",
            "PRODUCTO_4",
            "SKU_4",
            "Columna libre 13",
            "ASESOR DE VENTAS",
            "Columna libre 14",
            "Columna libre 15",
            "Columna libre 16",
            "Columna libre 17",
            "Columna libre 18",
            "Columna libre 19",
            "MARCA_1",
            "MODELO PRODUCTO_1",
            "Columna libre 20",
            "Columna libre 21",
            "Tipo Producto",
            "TIPO DE VALIDACION RENIEC",
            "CATEG FINAL",
            "ProductoSeguro"
        ]
        
        self.columnas_grupo1 = []
        self.columnas_grupo2 = []

    def _seleccionar_archivo(self, archivo_arg: str = None) -> str:
        """Abre diálogo para seleccionar archivo Excel o usa el argumento proporcionado"""
        if archivo_arg and os.path.exists(archivo_arg):
            return archivo_arg
        
        # Si no hay argumento o no existe, usar diálogo
        try:
            root = tk.Tk()
            root.withdraw()
            archivo = filedialog.askopenfilename(
                title="Seleccione archivo de ventas",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialdir=os.path.expanduser("~")
            )
            return archivo
        except Exception as e:
            logger.error(f"No se pudo abrir diálogo de selección: {e}")
            logger.info("Proporcione la ruta del archivo como argumento: python 01.2.RestructuraFNB.py <ruta_archivo>")
            return None

    def _parsear_fecha_dd_mm_yyyy(self, fecha_str):
        """Parsea fechas en formato dd/mm/yyyy"""
        if pd.isna(fecha_str) or fecha_str == '':
            return pd.NaT

        try:
            fecha_str = str(fecha_str).strip()
            if isinstance(fecha_str, pd.Timestamp):
                return fecha_str

            formatos = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d']

            for formato in formatos:
                try:
                    return pd.to_datetime(fecha_str, format=formato)
                except:
                    continue

            return pd.to_datetime(fecha_str, dayfirst=True, errors='coerce')

        except Exception as e:
            logger.warning(f"Error parseando fecha '{fecha_str}': {e}")
            return pd.NaT

    def _determinar_columnas_dinamicas(self, columnas_archivo: List[str]):
        """Determina qué columnas son grupo 1 y grupo 2"""
        # Grupo 2: columnas de productos que existen en el archivo
        self.columnas_grupo2 = [col for col in self.columnas_grupo2_fijas if col in columnas_archivo]
        # Grupo 1: todas las columnas restantes (incluye ESTADO, FECHA ENTREGA, etc)
        self.columnas_grupo1 = [col for col in columnas_archivo if col not in self.columnas_grupo2]
        logger.info(f"Columnas Grupo 1 (datos principales): {self.columnas_grupo1}")
        logger.info(f"Columnas Grupo 2 (productos): {self.columnas_grupo2}")

    def _generar_codigo_unico_vectorizado(self, df: pd.DataFrame) -> pd.Series:
        """Genera código único para cada fila basado en columnas de grupo 1"""
        return pd.util.hash_pandas_object(df[self.columnas_grupo1].astype(str).fillna(''), index=False)

    def _crear_mapeo_codigos(self, codigos_unicos: pd.Series) -> Dict[int, str]:
        """Crea mapeo de códigos únicos a códigos numerados"""
        unique_codes = codigos_unicos.unique()
        return {codigo: f"C{i + 1:07d}" for i, codigo in enumerate(unique_codes)}

    def _crear_columnas_pivotadas_ordenadas(self, df_g2_pivot: pd.DataFrame) -> pd.DataFrame:
        """Reordena las columnas pivotadas para que estén agrupadas por producto"""
        columnas_base = ['codigo_numerado']

        # Obtener todas las columnas pivotadas
        columnas_pivotadas = [col for col in df_g2_pivot.columns if col != 'codigo_numerado']

        # Crear lista ordenada de columnas finales
        columnas_finales = columnas_base.copy()

        # Agrupar por número de producto
        for i in range(1, self.max_productos + 1):
            for col_base in self.columnas_grupo2_fijas:
                col_nombre = f"{col_base}_{i}"
                if col_nombre in columnas_pivotadas:
                    columnas_finales.append(col_nombre)

        # Asegurar que todas las columnas estén presentes
        for col in columnas_pivotadas:
            if col not in columnas_finales:
                columnas_finales.append(col)

        return df_g2_pivot[columnas_finales]

    def _determinar_tipo_validacion_reniec(self, valor_origen: str, responsable: str) -> str:
        """Determina el tipo de validación RENIEC. Primero usa el origen, luego aplica la condición"""
        # Si hay valor en el origen, usarlo
        if pd.notna(valor_origen) and str(valor_origen).strip():
            return str(valor_origen).strip()
        
        # Si no hay valor en el origen, usar la condición del responsable
        if pd.isna(responsable):
            return ""
        
        responsable_upper = str(responsable).strip().upper()
        
        if responsable_upper == "CREDIWEB":
            return "RENIEC"
        elif responsable_upper == "CHATBOT":
            return "BIOMETRÍA"
        else:
            return ""

    def _aplicar_formato_excel(self, archivo_salida: str):
        """Aplica formato específico al archivo Excel"""
        try:
            wb = load_workbook(archivo_salida)
            ws = wb.active

            # Aplicar formato de fecha a columnas de fecha
            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and 'FECHA' in str(col_name.value).upper():
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].number_format = 'DD/MM/YYYY'
                        except:
                            pass

            # Aplicar formato de hora
            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and 'HORA' in str(col_name.value).upper():
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].number_format = '@'
                        except:
                            pass

            # Aplicar formato numérico a importes
            for col_idx, col_name in enumerate(ws[1], 1):
                if col_name.value and col_name.value in ["IMPORTE (S./)", "CRÉDITO UTILIZADO"]:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        try:
                            ws[f'{col_letter}{row}'].number_format = '#,##0.00'
                        except:
                            pass

            # Aplicar fuente Aptos tamaño 8 a todas las celdas
            font_aptos = Font(name='Aptos', size=8)

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font_aptos

            # Ajustar altura de filas
            for row_num in range(1, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 11.25

            # Autoajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(archivo_salida)
            logger.info("Formato aplicado correctamente al archivo Excel")

        except Exception as e:
            logger.error(f"Error aplicando formato Excel: {e}")

    def procesar(self, archivo_entrada: str = None):
        try:
            # Seleccionar archivo (usar argumento si se proporciona)
            archivo = self._seleccionar_archivo(archivo_entrada)
            if not archivo:
                logger.error("No se seleccionó archivo")
                return

            logger.info(f"Archivo seleccionado: {archivo}")

            # Cargar datos
            df = pd.read_excel(archivo)
            
            if df.empty:
                raise ValueError("Archivo sin datos")

            logger.info(f"Datos cargados: {len(df):,} filas")

            # Obtener columnas disponibles
            columnas_disponibles = df.columns.tolist()
            logger.info(f"Columnas encontradas: {len(columnas_disponibles)}")

            # Identificar columnas de fecha
            columnas_fecha = [col for col in columnas_disponibles if 'FECHA' in col.upper()]
            columnas_hora = [col for col in columnas_disponibles if 'HORA' in col.upper()]

            # Parsear fechas
            for fecha_col in columnas_fecha:
                if fecha_col in df.columns:
                    logger.info(f"Parseando columna de fecha: {fecha_col}")
                    df[fecha_col] = df[fecha_col].apply(self._parsear_fecha_dd_mm_yyyy)

            # Parsear horas
            for hora_col in columnas_hora:
                if hora_col in df.columns:
                    df[hora_col] = pd.to_datetime(df[hora_col], errors='coerce')

            # Determinar columnas dinámicas (grupo 1 y grupo 2)
            self._determinar_columnas_dinamicas(columnas_disponibles)

            # Generar códigos únicos basados en grupo 1
            logger.info(f"Total de filas antes de generar códigos: {len(df):,}")
            df['codigo_unico'] = self._generar_codigo_unico_vectorizado(df)
            logger.info(f"Códigos únicos generados: {df['codigo_unico'].nunique():,}")
            dict_codigos = self._crear_mapeo_codigos(df['codigo_unico'])
            df['codigo_numerado'] = df['codigo_unico'].map(dict_codigos)

            # Crear dataframe grupo 1 (sin duplicados - una fila por registro único)
            df_g1 = df.drop_duplicates('codigo_unico')[['codigo_numerado'] + self.columnas_grupo1].copy()
            logger.info(f"Registros únicos en Grupo 1 (df_g1): {len(df_g1):,}")

            # Crear dataframe grupo 2 (productos) - se repetirá por cada producto
            if self.columnas_grupo2:
                df_g2 = df[['codigo_unico'] + self.columnas_grupo2].copy()
                logger.info(f"Registros en Grupo 2 (df_g2) antes de filtrar: {len(df_g2):,}")
                df_g2['codigo_numerado'] = df_g2['codigo_unico'].map(dict_codigos)
                
                # Convertir PRECIO a numérico para ordenar correctamente
                df_g2['PRECIO'] = pd.to_numeric(df_g2['PRECIO'], errors='coerce').fillna(0)
                
                # ORDENAR por PRECIO descendente dentro de cada grupo antes de numerar
                df_g2 = df_g2.sort_values(['codigo_unico', 'PRECIO'], ascending=[True, False])
                
                # Ahora numerar productos después de ordenar por precio
                df_g2['producto_idx'] = df_g2.groupby('codigo_unico').cumcount() + 1
                logger.info(f"Distribución de producto_idx:\n{df_g2['producto_idx'].value_counts().sort_index()}")
                
                # Quedarse solo con los 4 productos de mayor precio
                df_g2 = df_g2[df_g2['producto_idx'] <= self.max_productos]
                logger.info(f"Registros en Grupo 2 después de filtrar (<=4): {len(df_g2):,}")

                # Pivotar datos de productos (melt + pivot)
                df_g2_long = pd.melt(df_g2, id_vars=['codigo_numerado', 'producto_idx'],
                                     value_vars=self.columnas_grupo2, var_name='atributo', value_name='valor')
                logger.info(f"Registros después de melt: {len(df_g2_long):,}")
                df_g2_long['col_final'] = df_g2_long['atributo'] + '_' + df_g2_long['producto_idx'].astype(str)
                df_g2_pivot = df_g2_long.pivot(index='codigo_numerado', columns='col_final', values='valor').reset_index()
                logger.info(f"Registros después de pivot (df_g2_pivot): {len(df_g2_pivot):,}")
                logger.info(f"Códigos únicos en df_g2_pivot: {df_g2_pivot['codigo_numerado'].nunique():,}")

                # Reordenar columnas pivotadas
                df_g2_pivot = self._crear_columnas_pivotadas_ordenadas(df_g2_pivot)

                # Unir dataframes
                logger.info(f"ANTES del merge - df_g1: {len(df_g1)} filas, df_g2_pivot: {len(df_g2_pivot)} filas")
                logger.info(f"Códigos únicos - df_g1: {df_g1['codigo_numerado'].nunique()}, df_g2_pivot: {df_g2_pivot['codigo_numerado'].nunique()}")
                df_final = df_g1.merge(df_g2_pivot, on='codigo_numerado', how='left')
                logger.info(f"DESPUÉS del merge - df_final: {len(df_final):,} filas")
                logger.info(f"Códigos únicos en df_final: {df_final['codigo_numerado'].nunique():,}")
                
                # Verificar si hay columnas duplicadas con sufijos _x y _y (esto indicaría un problema)
                cols_x = [col for col in df_final.columns if col.endswith('_x')]
                cols_y = [col for col in df_final.columns if col.endswith('_y')]
                if cols_x or cols_y:
                    logger.warning(f"Columnas duplicadas encontradas: {cols_x + cols_y}")
                    # Eliminar las versiones _y (mantener _x) y renombrar _x a sin sufijo
                    for col in cols_y:
                        if col in df_final.columns:
                            df_final = df_final.drop(columns=[col])
                    for col in cols_x:
                        if col in df_final.columns:
                            col_limpio = col.replace('_x', '')
                            df_final = df_final.rename(columns={col: col_limpio})
            else:
                # Si no hay columnas de grupo 2, usar solo grupo 1
                df_final = df_g1.copy()

            # Crear columnas de productos si faltan (para tener siempre 4)
            for i in range(1, self.max_productos + 1):
                for col_base in self.columnas_grupo2_fijas:
                    col_nombre = f"{col_base}_{i}"
                    if col_nombre not in df_final.columns:
                        df_final[col_nombre] = ""
                        logger.info(f"Columna creada (vacía): {col_nombre}")

            # Agregar columna TIPO DE VALIDACION RENIEC
            if 'RESPONSABLE DE VENTA' in df_final.columns:
                # Obtener valores del origen si existen
                if 'TIPO DE VALIDACION RENIEC' in df.columns:
                    # Crear mapeo de codigo_numerado a TIPO DE VALIDACION RENIEC del origen
                    mapeo_reniec_origen = df.drop_duplicates('codigo_unico')[['codigo_numerado', 'TIPO DE VALIDACION RENIEC']].set_index('codigo_numerado')['TIPO DE VALIDACION RENIEC'].to_dict()
                else:
                    mapeo_reniec_origen = {}
                
                df_final['TIPO DE VALIDACION RENIEC'] = df_final.apply(
                    lambda row: self._determinar_tipo_validacion_reniec(
                        mapeo_reniec_origen.get(row['codigo_numerado'], ""),
                        row['RESPONSABLE DE VENTA']
                    ),
                    axis=1
                )
                logger.info("Columna 'TIPO DE VALIDACION RENIEC' agregada")

            # Agregar columna Tipo Producto
            if 'PRODUCTO_1' in df_final.columns:
                producto_upper = df_final['PRODUCTO_1'].astype(str).str.upper()
                tiene_construccion = producto_upper.str.contains("PUNTO|DUCTE|ADICIONAL", na=False)
                tiene_multipunto = producto_upper.str.contains("MULTIPUNTO", na=False)
                df_final['Tipo Producto'] = np.where(
                    tiene_construccion & ~tiene_multipunto, 
                    "CON CONSTRUCCIÓN", 
                    "PRODUCTO SOLO"
                )
                logger.info("Columna 'Tipo Producto' agregada")
            else:
                df_final['Tipo Producto'] = ""

            # Agregar columna CATEG FINAL con lógica de fallback
            # Si CATEGORIA REAL está en blanco, usar CATEGORIA, sino usar CATEGORIA REAL
            df_final['CATEG FINAL'] = ''
            if 'CATEGORIA REAL' in df.columns and 'CATEGORIA' in df.columns:
                # Crear mapeos de codigo_numerado a ambas columnas
                mapeo_cat_real = df.drop_duplicates('codigo_unico')[['codigo_numerado', 'CATEGORIA REAL']].set_index('codigo_numerado')['CATEGORIA REAL'].to_dict()
                mapeo_cat = df.drop_duplicates('codigo_unico')[['codigo_numerado', 'CATEGORIA']].set_index('codigo_numerado')['CATEGORIA'].to_dict()
                
                # Aplicar la lógica: si CATEGORIA REAL está vacío, usar CATEGORIA
                def obtener_categ_final(row):
                    codigo_num = row['codigo_numerado']
                    cat_real = mapeo_cat_real.get(codigo_num, '')
                    if pd.isna(cat_real) or str(cat_real).strip() == '':
                        return mapeo_cat.get(codigo_num, '')
                    return cat_real
                
                df_final['CATEG FINAL'] = df_final.apply(obtener_categ_final, axis=1)
                logger.info("Lógica de CATEG FINAL aplicada (CATEGORIA REAL vacío → CATEGORIA)")
            elif 'CATEGORIA REAL' in df.columns:
                # Si solo existe CATEGORIA REAL
                mapeo_cat_real = df.drop_duplicates('codigo_unico')[['codigo_numerado', 'CATEGORIA REAL']].set_index('codigo_numerado')['CATEGORIA REAL'].to_dict()
                df_final['CATEG FINAL'] = df_final['codigo_numerado'].map(mapeo_cat_real)
            elif 'CATEGORIA' in df.columns:
                # Si solo existe CATEGORIA
                mapeo_cat = df.drop_duplicates('codigo_unico')[['codigo_numerado', 'CATEGORIA']].set_index('codigo_numerado')['CATEGORIA'].to_dict()
                df_final['CATEG FINAL'] = df_final['codigo_numerado'].map(mapeo_cat)

            # Agregar columna ProductoSeguro
            if 'ALIADO COMERCIAL' in df_final.columns:
                aliado_upper = df_final['ALIADO COMERCIAL'].astype(str).str.strip().str.upper()
                df_final['ProductoSeguro'] = np.where(aliado_upper == 'CARDIF', 'SEGURO', 'PRODUCTO')
            else:
                df_final['ProductoSeguro'] = ''

            # Crear columnas vacías (blancas)
            for col_libre in self.columnas_finales_orden:
                if col_libre.startswith("Columna libre") and col_libre not in df_final.columns:
                    df_final[col_libre] = ""

            # Reordenar columnas según el orden especificado
            columnas_a_usar = []
            for col in self.columnas_finales_orden:
                if col in df_final.columns:
                    columnas_a_usar.append(col)

            # Eliminar columnas auxiliares (codigo_numerado no está en orden final, se eliminará)
            df_final = df_final[columnas_a_usar]
            
            # Asegurar que codigo_numerado se elimine si aún existe
            if 'codigo_numerado' in df_final.columns:
                df_final = df_final.drop(columns=['codigo_numerado'])

            # FILTRAR POR ESTADO: Solo ENTREGADO, PENDIENTE DE ENTREGA, PENDIENTE DE APROBACIÓN
            if 'ESTADO' in df_final.columns:
                estados_validos = ['ENTREGADO', 'PENDIENTE DE ENTREGA', 'PENDIENTE DE APROBACIÓN']
                df_final['ESTADO'] = df_final['ESTADO'].astype(str).str.strip().str.upper()
                registros_antes_filtro = len(df_final)
                df_final = df_final[df_final['ESTADO'].isin(estados_validos)].copy()
                registros_despues_filtro = len(df_final)
                logger.info(f"Registros antes de filtro ESTADO: {registros_antes_filtro:,}")
                logger.info(f"Registros después de filtro ESTADO: {registros_despues_filtro:,}")
                logger.info(f"Registros eliminados: {registros_antes_filtro - registros_despues_filtro:,}")
                
                # Reemplazar PENDIENTE DE APROBACIÓN por PENDIENTE DE ENTREGA
                df_final['ESTADO'] = df_final['ESTADO'].replace('PENDIENTE DE APROBACIÓN', 'PENDIENTE DE ENTREGA')
                logger.info("Valores 'PENDIENTE DE APROBACIÓN' reemplazados por 'PENDIENTE DE ENTREGA'")

            # Ordenar por FECHA VENTA y HORA VENTA
            columnas_ordenamiento = []
            if 'FECHA VENTA' in df_final.columns:
                columnas_ordenamiento.append('FECHA VENTA')
            if 'HORA VENTA' in df_final.columns:
                columnas_ordenamiento.append('HORA VENTA')
            
            if columnas_ordenamiento:
                df_final = df_final.sort_values(columnas_ordenamiento).reset_index(drop=True)
                logger.info(f"Datos ordenados por: {', '.join(columnas_ordenamiento)}")

            # Formatear columnas numéricas
            for col in ["IMPORTE (S./)", "CRÉDITO UTILIZADO"]:
                if col in df_final.columns:
                    df_final[col] = pd.to_numeric(df_final[col], errors='coerce').round(2)

            # Formatear HORA VENTA
            if 'HORA VENTA' in df_final.columns:
                hora_temp = pd.to_datetime(df_final['HORA VENTA'], errors='coerce')
                df_final['HORA VENTA'] = hora_temp.dt.strftime('%H:%M')
                df_final['HORA VENTA'] = df_final['HORA VENTA'].fillna('')

            # Guardar archivo en la misma ruta que el origen
            ruta_origen = os.path.dirname(archivo)
            nombre_archivo = "Archivo_Reestructurado.xlsx"
            ruta_salida = os.path.join(ruta_origen, nombre_archivo)

            # Guardar sin formato primero
            df_final.to_excel(ruta_salida, index=False)

            # Aplicar formato
            self._aplicar_formato_excel(ruta_salida)

            logger.info(f"Archivo generado: {ruta_salida}")
            print(f"EXITO:{len(df_final)}")  # Mensaje simple para el launcher

        except Exception as e:
            logger.error(f"Error en procesamiento: {e}")
            import traceback
            traceback.print_exc()
            print(f"ERROR:{str(e)}")  # Mensaje simple de error para el launcher


if __name__ == "__main__":
    restructurer = SalesDataRestructurer()
    
    # Verificar si se proporcionó un archivo como argumento
    archivo_entrada = None
    if len(sys.argv) > 1:
        archivo_entrada = sys.argv[1]
        if not os.path.exists(archivo_entrada):
            print(f"❌ ERROR: El archivo no existe: {archivo_entrada}")
            sys.exit(1)
    
    restructurer.procesar(archivo_entrada)
